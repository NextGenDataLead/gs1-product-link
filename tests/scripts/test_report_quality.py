"""Tests for scripts/report_quality.py — the CLI that loads the issue files and writes the report.

Drives ``main`` in a temp working directory: writes sample ``output/{client}/data/*_issues.json``
(and products.json), then asserts the consolidated markdown report is written and the exit code.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from lib.records import LocalisedText, ProductRecord
from scripts import report_quality


def _write(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload), encoding="utf-8")


def _write_process_list(tmp_path: Path, gtins: list[str]) -> None:
    """A process list at the path `clients.yml` names, so `in_scope` really narrows."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="Barcode")
    for row, gtin in enumerate(gtins, start=2):
        sheet.cell(row=row, column=1, value=gtin)
    path = tmp_path / "input" / "noviplast" / "process-list.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _seed(tmp_path: Path) -> Path:
    data = tmp_path / "output" / "noviplast" / "data"
    _write(
        data / "generated_issues.json",
        [
            {
                "gtin": "08713195003276",
                "field": "description_short.nl",
                "source": "attr 1083",
                "issue": "missing_generation_input",
                "value": "",
                "detail": "d",
            },
            {
                "gtin": "08713195007915",
                "field": "generated_description.nl",
                "source": "inferred",
                "issue": "generation_inference",
                "value": "magnet claim",
                "detail": "d",
            },
        ],
    )
    _write(data / "source_issues.json", [])
    _write(data / "video_map_issues.json", [])
    _write(data / "category_issues.json", [])
    # Both GTINs the seeded findings name: a finding about a product the file does not carry is
    # not in scope, because scope is decided over the parsed products.
    _write(
        data / "products.json",
        [
            ProductRecord(
                gtin=gtin,
                brand="Noviplast",
                product_name=LocalisedText(values={"nl": name}),
            ).model_dump(mode="json")
            for gtin, name in (("08713195007915", "LED-lamp"), ("08713195003276", "plasma"))
        ],
    )
    return data


def test_writes_report_and_exit_0(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)
    _seed(tmp_path)

    code = report_quality.main(["noviplast"])

    assert code == 0
    report = tmp_path / "output" / "noviplast" / "data-quality-report.md"
    assert report.is_file()
    text = report.read_text(encoding="utf-8")
    assert "Data quality report" in text
    assert "08713195003276" in text  # held GTIN
    assert "magnet claim" in text  # inference surfaced


def test_findings_for_out_of_scope_gtins_never_reach_the_report(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The report describes one run, so every section describes the same set of GTINs.

    §3 was the exception: its rows came straight from `source_issues.json`, which the parser
    writes over the whole workbook, while §0's matrix and §1's rows were computed over the
    process list. Four of eleven GTINs in §3 were outside the run — findable nowhere else in the
    document, which is how it was noticed. `08713195000794` is in the export and not on the
    process list.
    """
    monkeypatch.chdir(tmp_path)
    data = _seed(tmp_path)
    in_scope_gtin, out_of_scope_gtin = "08713195007915", "08713195000794"
    # The real process list lives at a path relative to the repo root, so under `chdir(tmp_path)`
    # it is absent and `in_scope` narrows nothing — a scope test would pass on no scope at all.
    _write_process_list(tmp_path, [in_scope_gtin, "08713195003276"])
    _write(
        data / "source_issues.json",
        [
            {
                "gtin": gtin,
                "field": "net_content",
                "source": "TradeItemMeasurements attr 3510",
                "issue": "value_inconsistent_across_markets",
                "value": "x",
                "detail": "d",
                "market_values": [["528", "x"], ["056", "y"]],
            }
            for gtin in (in_scope_gtin, out_of_scope_gtin)
        ],
    )
    # Every issue file, not only the one where the leak was noticed. `generated_issues` happens
    # to be scoped already — generation only runs for in-scope units — so a narrowing applied to
    # `source` alone passes today and leaks the moment that stops being true.
    _write(
        data / "generated_issues.json",
        [
            {
                "gtin": out_of_scope_gtin,
                "field": "generated_description.nl",
                "source": "inferred",
                "issue": "generation_inference",
                "value": "a claim about a product this run will not touch",
                "detail": "d",
            }
        ],
    )
    # A finding with no GTIN is about the *input*, not a product, so scope cannot apply to it.
    _write(
        data / "video_map_issues.json",
        [
            {
                "gtin": "",
                "field": "video",
                "source": "mapping.yml",
                "issue": "video_unconfirmed",
                "value": "Aqua Mat v2.mp4",
                "detail": "d",
            }
        ],
    )
    products = json.loads((data / "products.json").read_text())
    products.append(
        ProductRecord(
            gtin=out_of_scope_gtin,
            brand="Noviplast",
            product_name=LocalisedText(values={"nl": "Power jet"}),
        ).model_dump(mode="json")
    )
    _write(data / "products.json", products)

    assert report_quality.main(["noviplast"]) == 0

    text = (tmp_path / "output" / "noviplast" / "data-quality-report.md").read_text()
    assert in_scope_gtin in text
    assert out_of_scope_gtin not in text
    assert "a claim about a product this run will not touch" not in text
    assert "Aqua Mat v2.mp4" in text  # kept: it names a video file, not a GTIN


def test_missing_data_dir_exits_2(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.chdir(tmp_path)  # no output/ at all

    assert report_quality.main(["noviplast"]) == 2


def test_absent_issue_files_treated_as_empty(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.chdir(tmp_path)
    # data dir exists but only products.json present; issue files absent
    data = tmp_path / "output" / "noviplast" / "data"
    data.mkdir(parents=True)
    (data / "products.json").write_text("[]", encoding="utf-8")

    assert report_quality.main(["noviplast"]) == 0
    assert (tmp_path / "output" / "noviplast" / "data-quality-report.md").is_file()


def test_generated_at_is_local_time_with_a_named_zone() -> None:
    """Local, so it matches the file browser beside the file; named, so it travels unambiguously."""
    stamp = report_quality._generated_at()

    # "2026-08-13 22:02 CEST" — date, time, then a zone abbreviation.
    assert re.fullmatch(r"\d{4}-\d{2}-\d{2} \d{2}:\d{2} \S+", stamp), stamp
    assert stamp.startswith(datetime.now().astimezone().strftime("%Y-%m-%d"))
