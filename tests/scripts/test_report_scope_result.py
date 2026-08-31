"""Tests for scripts/report_scope_result.py — the per-run result sheet.

Read-only orchestration over ``lib.scope_report``: drive ``main`` with a faked ``get_client`` and
a temp working directory, asserting where the file lands, what is in it, and the exit codes. The
join itself is covered in ``tests/lib/test_scope_report.py``.

Two behaviours here are the script's own and nothing else covers them: choosing the newest run by
**modification time** rather than by name, and refusing a ``plan.json`` that belongs to a later
run — which, for anything but the newest log, is what it usually is.
"""

from __future__ import annotations

import json
import os
import time
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import pytest

from lib.config import (
    ClientConfig,
    ExportConfig,
    GS1Config,
    ProcessListConfig,
    WordPressConfig,
)
from lib.records import (
    LocalisedText,
    Plan,
    ProductRecord,
    RunOutcome,
    SkippedUnit,
    SkipReason,
)
from scripts import report_scope_result

GTIN = "08713195000001"
MISSING = "08713195000002"


def _config(process_list: ProcessListConfig | None) -> ClientConfig:
    return ClientConfig(
        client_id="acme",
        display_name="Acme BV",
        gs1=GS1Config(
            account_number_test="8720796420906",
            client_id_env_test="GS1_CID",
            client_secret_env_test="GS1_SEC",
        ),
        export=ExportConfig(path="input/acme.xlsx"),
        wordpress=WordPressConfig(
            site_url="https://wp.test",
            username="bot",
            app_password_env="WP_PASS",
            languages=["nl", "fr"],
        ),
        process_list=process_list,
    )


def _write_list(path: Path, rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Artikelnr.", "Omschrijving NL", "Barcode"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def _outcome(gtin: str, language: str, status: str, **kwargs: object) -> str:
    return RunOutcome(
        gtin=gtin, language=language, ts=datetime(2026, 8, 27, tzinfo=UTC), status=status, **kwargs
    ).model_dump_json()


@pytest.fixture
def workspace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """A temp cwd with a scope list, a parsed export, and one run log."""
    monkeypatch.chdir(tmp_path)
    _write_list(
        tmp_path / "input" / "process-list.xlsx",
        [["1079", "Drain saver", "8713195000001"], ["3086", "Contour King", "8713195000002"]],
    )
    products = tmp_path / "output" / "acme" / "data" / "products.json"
    products.parent.mkdir(parents=True)
    products.write_text(
        json.dumps(
            [
                ProductRecord(
                    gtin=GTIN,
                    brand="Acme",
                    product_name=LocalisedText(values={"nl": "Drain saver"}),
                ).model_dump(mode="json")
            ]
        ),
        encoding="utf-8",
    )
    runs = tmp_path / "output" / "acme" / "runs"
    runs.mkdir()
    (runs / "20260827T085405Z.jsonl").write_text(
        _outcome(GTIN, "nl", "ok", wp_url="https://wp.test/nl/p-1", wp_page_id=1637)
        + "\n"
        + _outcome(GTIN, "fr", "error", error="500 boom", failed_call="POST /pages")
        + "\n",
        encoding="utf-8",
    )
    return tmp_path


def _patch_client(monkeypatch: pytest.MonkeyPatch, cfg: ClientConfig) -> None:
    monkeypatch.setattr(report_scope_result, "get_client", lambda _cid: cfg)


def _list_config(tmp_path: Path) -> ProcessListConfig:
    return ProcessListConfig(path=str(tmp_path / "input" / "process-list.xlsx"))


def _sheets(path: Path) -> dict[str, list[tuple[object, ...]]]:
    workbook = openpyxl.load_workbook(path)
    return {name: list(workbook[name].iter_rows(values_only=True)) for name in workbook.sheetnames}


def test_the_workbook_has_the_three_sheets_and_the_operators_header(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Arrange
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    code = report_scope_result.main(["acme"])

    # Assert
    assert code == 0
    out = workspace / "output" / "acme" / "runs" / "20260827T085405Z-scope.xlsx"
    sheets = _sheets(out)
    assert list(sheets) == ["scope", "units", "legend"]
    assert sheets["scope"][0] == (
        "Artikelnr.",
        "Omschrijving NL",
        "Barcode",
        "in_scope",
        "result",
        "status_nl",
        "page_nl",
        "detail_nl",
        "status_fr",
        "page_fr",
        "detail_fr",
    )


def test_one_known_row_reads_the_way_the_run_went(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Published in Dutch, failed in French — the case a single status column cannot hold."""
    # Arrange
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    report_scope_result.main(["acme"])

    # Assert
    scope = _sheets(workspace / "output" / "acme" / "runs" / "20260827T085405Z-scope.xlsx")["scope"]
    published, missing = scope[1], scope[2]
    assert published[:3] == ("1079", "Drain saver", "8713195000001")
    assert published[3:5] == ("yes", "error")
    assert published[5:8] == ("ok", "https://wp.test/nl/p-1", None)
    assert published[8] == "error"
    assert "POST /pages" in str(published[10])
    assert missing[3:5] == ("not in export", "not run"), (
        "a SKU the export has no row for is named as such, not reported as a failure"
    )


def test_the_units_sheet_carries_a_row_per_unit(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Arrange
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    report_scope_result.main(["acme"])

    # Assert
    units = _sheets(workspace / "output" / "acme" / "runs" / "20260827T085405Z-scope.xlsx")["units"]
    assert [(row[1], row[3]) for row in units[1:]] == [("nl", "ok"), ("fr", "error")]


def test_the_newest_run_is_chosen_by_mtime_not_by_name(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """``{ts}-1.jsonl`` sorts *before* ``{ts}.jsonl`` — ``-`` precedes ``.``.

    Two runs a second apart is not hypothetical: it is what a re-run after a failure looks like,
    and reporting on the wrong one of the pair is invisible until somebody opens a page URL.
    """
    # Arrange: the same-second second run, written later.
    runs = workspace / "output" / "acme" / "runs"
    later = runs / "20260827T085405Z-1.jsonl"
    later.write_text(_outcome(GTIN, "nl", "dry-run") + "\n", encoding="utf-8")
    now = time.time()
    os.utime(runs / "20260827T085405Z.jsonl", (now - 60, now - 60))
    os.utime(later, (now, now))
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    report_scope_result.main(["acme"])

    # Assert
    assert (runs / "20260827T085405Z-1-scope.xlsx").exists()
    assert not (runs / "20260827T085405Z-scope.xlsx").exists()


def test_the_uploaded_list_is_reported_so_a_deselected_row_still_appears(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The report is about the batch the operator asked for, not only the rows that ran."""
    # Arrange: the archive holds both rows; the control file holds one.
    _write_list(
        workspace / "input" / "process-list.source.xlsx",
        [["1079", "Drain saver", "8713195000001"], ["3086", "Contour King", "8713195000002"]],
    )
    _write_list(workspace / "input" / "process-list.xlsx", [["1079", "Drain saver", GTIN]])
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    report_scope_result.main(["acme"])

    # Assert
    scope = _sheets(workspace / "output" / "acme" / "runs" / "20260827T085405Z-scope.xlsx")["scope"]
    assert [row[3] for row in scope[1:]] == ["yes", "not selected"]


def test_a_plan_generated_after_the_run_is_refused(
    workspace: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    """``plan.json`` is overwritten by every ``run_plan``, so for an old run it is somebody else's.

    Reporting a later plan's holds as this run's would put ``held`` beside rows this run never
    saw — a wrong fact wearing the right word, which is the failure mode this whole report is
    designed against.
    """
    # Arrange
    plan = Plan(
        client_id="acme",
        generated_at=datetime(2026, 9, 1, tzinfo=UTC),
        total=0,
        counts={},
        rows=[],
        skipped=[
            SkippedUnit(
                gtin=MISSING,
                language="nl",
                reason=SkipReason.NO_CONFIRMED_VIDEO,
                detail="no video",
            )
        ],
    )
    (workspace / "output" / "acme" / "plan.json").write_text(
        plan.model_dump_json(), encoding="utf-8"
    )
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    report_scope_result.main(["acme"])

    # Assert
    assert "a later run's plan" in capsys.readouterr().err
    units = _sheets(workspace / "output" / "acme" / "runs" / "20260827T085405Z-scope.xlsx")["units"]
    assert all(row[2] == "run log" for row in units[1:]), "no hold from the wrong plan leaked in"


def test_a_plan_from_before_the_run_supplies_its_holds(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Arrange
    plan = Plan(
        client_id="acme",
        generated_at=datetime(2026, 8, 27, tzinfo=UTC),
        total=0,
        counts={},
        rows=[],
        skipped=[
            SkippedUnit(
                gtin=MISSING,
                language="nl",
                reason=SkipReason.MISSING_MANDATORY_FIELD,
                detail="dim_height is blank",
            )
        ],
    )
    (workspace / "output" / "acme" / "plan.json").write_text(
        plan.model_dump_json(), encoding="utf-8"
    )
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    report_scope_result.main(["acme"])

    # Assert
    units = _sheets(workspace / "output" / "acme" / "runs" / "20260827T085405Z-scope.xlsx")["units"]
    assert ("plan", "held") in [(row[2], row[3]) for row in units[1:]]


def test_no_run_log_is_a_usage_error_not_an_empty_report(
    workspace: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    """An empty workbook would read as "nothing went wrong". Nothing ran."""
    # Arrange
    for path in (workspace / "output" / "acme" / "runs").glob("*.jsonl"):
        path.unlink()
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act
    code = report_scope_result.main(["acme"])

    # Assert
    assert code == 2
    assert "nothing to report on yet" in capsys.readouterr().err


def test_a_missing_run_log_path_is_an_error(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Arrange
    _patch_client(monkeypatch, _config(_list_config(workspace)))

    # Act / Assert
    assert report_scope_result.main(["acme", "--run", str(workspace / "nope.jsonl")]) == 1


def test_a_client_with_no_scope_list_is_a_usage_error(
    workspace: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    # Arrange
    _patch_client(monkeypatch, _config(None))

    # Act
    code = report_scope_result.main(["acme"])

    # Assert
    assert code == 2
    assert "no process_list block" in capsys.readouterr().err


def test_out_overrides_where_the_workbook_lands(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Arrange
    _patch_client(monkeypatch, _config(_list_config(workspace)))
    out = workspace / "elsewhere" / "result.xlsx"

    # Act
    code = report_scope_result.main(["acme", "--out", str(out)])

    # Assert
    assert code == 0
    assert out.is_file()
