"""Tests for scripts/parse_export.py (IMPLEMENTATION_SPEC §8.1)."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
import pytest
import yaml

from lib.config import get_client
from scripts import parse_export


def _write_flat_xlsx(tmp_path: Path, header: list[str], rows: list[list[Any]]) -> str:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    path = tmp_path / "flat.xlsx"
    workbook.save(path)
    return str(path)


def _patch_client(monkeypatch: pytest.MonkeyPatch, tmp_path: Path, export: dict[str, Any]) -> None:
    config = {
        "version": 1,
        "clients": {
            "acme": {
                "display_name": "Acme",
                "gs1": {
                    "account_number_test": "8720796420906",
                    "client_id_env_test": "ID",
                    "client_secret_env_test": "SECRET",
                },
                "export": export,
                "wordpress": {
                    "site_url": "https://acme.test",
                    "username": "bot",
                    "app_password_env": "WP",
                    "default_language": "nl",
                },
            }
        },
    }
    cfg_path = tmp_path / "clients.yml"
    cfg_path.write_text(yaml.safe_dump(config), encoding="utf-8")
    monkeypatch.setattr(parse_export, "get_client", lambda cid: get_client(cid, cfg_path))


_FLAT_MAP = {"GTIN": "gtin", "Merk": "brand", "Productnaam NL": "product_name.nl"}


def test_flat_happy_path_writes_output(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xlsx = _write_flat_xlsx(
        tmp_path,
        ["GTIN", "Merk", "Productnaam NL"],
        [["08713195007359", "Noviplast", "Rugsteun"], ["05031694050403", "Noviplast", "Pin"]],
    )
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})
    out = tmp_path / "products.json"

    code = parse_export.main(["acme", "--output", str(out)])

    assert code == 0
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert {p["gtin"] for p in payload} == {"08713195007359", "05031694050403"}


def test_duplicate_gtin_first_wins_with_warning(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    # E3: a repeated GTIN keeps the first occurrence; the rest warn and are skipped.
    xlsx = _write_flat_xlsx(
        tmp_path,
        ["GTIN", "Merk", "Productnaam NL"],
        [
            ["08713195007359", "Noviplast", "Rugsteun"],
            ["08713195007359", "Noviplast", "Duplicaat"],
        ],
    )
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})
    out = tmp_path / "products.json"

    code = parse_export.main(["acme", "--output", str(out)])

    assert code == 0
    assert "(1 warnings)" in capsys.readouterr().err
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert len(payload) == 1
    assert payload[0]["product_name"]["values"]["nl"] == "Rugsteun"


def test_empty_rows_skipped_silently(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    # E4: a blank row is skipped without a warning.
    xlsx = _write_flat_xlsx(
        tmp_path,
        ["GTIN", "Merk", "Productnaam NL"],
        [["08713195007359", "Noviplast", "Rugsteun"], [None, None, None]],
    )
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})

    code = parse_export.main(["acme", "--dry-run"])

    assert code == 0


def test_unmapped_column_warns(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    # E16: a column present but not in column_map/extras_columns is a warning.
    xlsx = _write_flat_xlsx(
        tmp_path,
        ["GTIN", "Merk", "Productnaam NL", "Surplus"],
        [["08713195007359", "Noviplast", "Rugsteun", "x"]],
    )
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})

    code = parse_export.main(["acme", "--dry-run"])

    assert code == 0
    assert "(1 warnings)" in capsys.readouterr().err


def test_missing_required_column_is_parse_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # E17 (required): a required target's column is absent from the sheet.
    xlsx = _write_flat_xlsx(tmp_path, ["GTIN", "Merk"], [["08713195007359", "Noviplast"]])
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})

    assert parse_export.main(["acme", "--dry-run"]) == 1


def test_missing_optional_column_warns(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    # E17 (optional): a mapped extras column absent from the sheet warns but succeeds.
    xlsx = _write_flat_xlsx(
        tmp_path,
        ["GTIN", "Merk", "Productnaam NL"],
        [["08713195007359", "Noviplast", "Rugsteun"]],
    )
    export = {"path": xlsx, "column_map": _FLAT_MAP, "extras_columns": ["HS-code"]}
    _patch_client(monkeypatch, tmp_path, export)

    code = parse_export.main(["acme", "--dry-run"])

    assert code == 0
    assert "(1 warnings)" in capsys.readouterr().err


def test_dry_run_writes_no_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xlsx = _write_flat_xlsx(
        tmp_path, ["GTIN", "Merk", "Productnaam NL"], [["08713195007359", "Noviplast", "R"]]
    )
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})
    out = tmp_path / "products.json"

    parse_export.main(["acme", "--dry-run", "--output", str(out)])

    assert not out.exists()


def test_config_error_returns_exit_2(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xlsx = _write_flat_xlsx(
        tmp_path, ["GTIN", "Merk", "Productnaam NL"], [["08713195007359", "Noviplast", "R"]]
    )
    _patch_client(monkeypatch, tmp_path, {"path": xlsx, "column_map": _FLAT_MAP})

    assert parse_export.main(["unknown-client", "--dry-run"]) == 2


def test_gdsn_integration_writes_output(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xlsx = _write_gdsn_xlsx(tmp_path)
    _patch_client(monkeypatch, tmp_path, _gdsn_export(xlsx))
    monkeypatch.chdir(tmp_path)
    out = tmp_path / "products.json"

    code = parse_export.main(["acme", "--output", str(out)])

    assert code == 0
    payload = json.loads(out.read_text(encoding="utf-8"))
    assert payload[0]["product_name"]["values"] == {"nl": "Rugsteun NL", "fr": "Support FR"}


def _gdsn_export(xlsx: str, *, strip_prefix: str = "") -> dict[str, object]:
    name_src: dict[str, object] = {
        "sheet": "TradeItemDescription",
        "attribute": "3297",
        "localised": True,
    }
    if strip_prefix:
        name_src["strip_prefix"] = strip_prefix
    return {
        "format": "gdsn",
        "path": xlsx,
        "market_language": {"528": "nl", "056": "fr"},
        "gdsn_map": {
            "product_name": name_src,
            "brand": {"sheet": "TradeItemDescription", "attribute": "3336"},
        },
    }


def test_source_issues_report_is_written_with_the_findings(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    """Datapool defects must land in a file, not just a log line.

    The operator fixes these later, in MyGS1, from a work queue — a warning that scrolled
    past in a terminal is not one. "Rugsteun NL" against strip_prefix "Rugsteun_" is a
    near-miss, so it reports rather than strips.
    """
    xlsx = _write_gdsn_xlsx(tmp_path)
    _patch_client(monkeypatch, tmp_path, _gdsn_export(xlsx, strip_prefix="Rugsteun_"))
    monkeypatch.chdir(tmp_path)

    code = parse_export.main(["acme", "--output", str(tmp_path / "products.json")])

    assert code == 0  # non-fatal: the products still parse
    report = tmp_path / "output" / "acme" / "data" / "source_issues.json"
    issues = json.loads(report.read_text(encoding="utf-8"))
    assert len(issues) == 1
    assert issues[0] == {
        "gtin": "08713195007359",
        "field": "product_name.nl",
        # The operator searches MyGS1 for *this*, not for our "product_name".
        "source": "TradeItemDescription attr 3297",
        "issue": "brand_prefix_mismatch",
        "value": "Rugsteun NL",
        "detail": issues[0]["detail"],  # wording asserted below
    }
    assert "resembles but does not match" in issues[0]["detail"]
    # And the operator is pointed at the file, with the count.
    err = capsys.readouterr().err
    assert "1 source-data issue(s) need fixing at the source" in err
    assert "source_issues.json" in err


def test_source_issues_report_written_even_when_clean(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    """An empty report means "checked, nothing found"; a missing one means "never looked"."""
    xlsx = _write_gdsn_xlsx(tmp_path)
    _patch_client(monkeypatch, tmp_path, _gdsn_export(xlsx))
    monkeypatch.chdir(tmp_path)

    parse_export.main(["acme", "--output", str(tmp_path / "products.json")])

    report = tmp_path / "output" / "acme" / "data" / "source_issues.json"
    assert json.loads(report.read_text(encoding="utf-8")) == []
    assert "source-data issue" not in capsys.readouterr().err  # nothing to shout about


def test_dry_run_writes_no_issues_report(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    xlsx = _write_gdsn_xlsx(tmp_path)
    _patch_client(monkeypatch, tmp_path, _gdsn_export(xlsx, strip_prefix="Rugsteun_"))
    monkeypatch.chdir(tmp_path)

    parse_export.main(["acme", "--dry-run"])

    assert not (tmp_path / "output" / "acme" / "data" / "source_issues.json").exists()


def _write_gdsn_xlsx(tmp_path: Path) -> str:
    header = [
        [
            "Gtin",
            "TargetMarketCountryCode",
            "InformationProviderOfTradeItem",
            "TradeItemUnitDescriptorCode",
            "TradeItemDescriptionInformation",
            "TradeItemDescriptionInformation",
            "TradeItemDescriptionInformation",
        ],
        [
            None,
            None,
            None,
            None,
            "DescriptionShort[0]",
            "DescriptionShort[0]",
            "BrandNameInformation",
        ],
        [None, None, None, None, "LanguageCode", "Value", "BrandName"],
        [None] * 7,
        [None] * 7,
        [None] * 7,
        [
            "GTIN (3059)",
            "Country (3179)",
            "Provider (3088)",
            "Unit (3074)",
            "Short product name (3297)",
            "Short product name (3297)",
            "Brand Name (3336)",
        ],
    ]
    data = [
        ["08713195007359", "528", "GLN", "BASE_UNIT_OR_EACH", "nl", "Rugsteun NL", "Noviplast"],
        ["08713195007359", "056", "GLN", "BASE_UNIT_OR_EACH", "fr", "Support FR", "Noviplast"],
    ]
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    sheet = workbook.create_sheet("TradeItemDescription")
    for row in [*header, *data]:
        sheet.append(row)
    path = tmp_path / "gdsn.xlsx"
    workbook.save(path)
    return str(path)
