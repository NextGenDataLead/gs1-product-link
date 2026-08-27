"""Tests for scripts/report_video_candidates.py (issue #120).

Read-only orchestration over ``lib.video_candidates``: drive ``main`` with a faked ``get_client``
and a temp working directory, asserting where the file lands, what each format contains, and the
exit codes. Row building itself is covered in ``tests/lib/test_video_candidates.py``.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
import pytest

from lib.config import (
    ClientConfig,
    ExportConfig,
    GS1Config,
    MediaConfig,
    WordPressConfig,
)
from lib.records import LocalisedText, ProductRecord
from scripts import report_video_candidates

_MAPPING = """nl:
  - {file: "DrainSticks_NL.mpeg", gtin: ""}
  - {file: "Bulbman.mpg", gtin: "8713195007434"}
fr:
  - {file: "DrainSticks_FR.mpeg", gtin: ""}
"""


def _make_config(media: MediaConfig | None) -> ClientConfig:
    return ClientConfig(
        client_id="acme",
        display_name="Acme BV",
        gs1=GS1Config(
            account_number_test="8720796420906",
            client_id_env_test="GS1_CID",
            client_secret_env_test="GS1_SEC",
        ),
        export=ExportConfig(path="input/acme.xlsx"),
        wordpress=WordPressConfig(
            site_url="https://wp.test",
            username="bot",
            app_password_env="WP_PASS",
            languages=["nl", "fr"],
        ),
        media=media,
    )


def _products() -> list[ProductRecord]:
    return [
        ProductRecord(
            gtin="08713195007434",
            brand="Noviplast",
            product_name=LocalisedText(values={"nl": "stickylamp", "fr": "Lampe portable"}),
            extras_localised={
                "logistics_name": LocalisedText(values={"nl": "Bulb man", "fr": "Bulb man"})
            },
        ),
        ProductRecord(
            gtin="08713195006796",
            brand="Noviplast",
            product_name=LocalisedText(values={"nl": "afvoerreinigingsstick"}),
            extras_localised={"logistics_name": LocalisedText(values={"fr": "Drain Sticks 12pc"})},
        ),
    ]


@pytest.fixture
def workspace(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    """A temp cwd holding both video folders, the mapping, and ``products.json``."""
    monkeypatch.chdir(tmp_path)
    for language, names in {
        "nl": ["DrainSticks_NL.mpeg", "Bulbman.mpg", "Stray.mpg"],
        "fr": ["DrainSticks_FR.mpeg"],
    }.items():
        folder = tmp_path / language
        folder.mkdir()
        for name in names:
            (folder / name).write_bytes(b"x")
    (tmp_path / "mapping.yml").write_text(_MAPPING, encoding="utf-8")
    products = tmp_path / "output" / "acme" / "data" / "products.json"
    products.parent.mkdir(parents=True)
    products.write_text(
        json.dumps([p.model_dump(mode="json") for p in _products()]), encoding="utf-8"
    )
    return tmp_path


def _media(tmp_path: Path, *, map_path: str | None = "mapping.yml") -> MediaConfig:
    return MediaConfig(
        video_folders={"nl": str(tmp_path / "nl"), "fr": str(tmp_path / "fr")},
        video_map_path=map_path,
    )


def _patch_client(monkeypatch: pytest.MonkeyPatch, cfg: ClientConfig) -> None:
    monkeypatch.setattr(report_video_candidates, "get_client", lambda _cid: cfg)


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


# --- Where it lands ----------------------------------------------------------


def test_the_default_output_is_a_spreadsheet_beside_the_other_derived_artefacts(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _patch_client(monkeypatch, _make_config(_media(workspace)))

    code = report_video_candidates.main(["acme"])

    assert code == 0
    assert (workspace / "output" / "acme" / "video-map-candidates.xlsx").is_file()


def test_csv_is_written_with_a_bom_so_excel_does_not_mangle_the_french_names(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Without it Excel on Windows opens a UTF-8 CSV as cp1252, and every accented name in this
    feed arrives broken — in the file that is sent to the client."""
    _patch_client(monkeypatch, _make_config(_media(workspace)))
    out = workspace / "report.csv"

    assert report_video_candidates.main(["acme", "--format", "csv", "--out", str(out)]) == 0
    assert out.read_bytes().startswith(b"\xef\xbb\xbf")


# --- What is in it -----------------------------------------------------------


def test_the_rows_are_the_union_of_the_mapping_and_the_folders(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _patch_client(monkeypatch, _make_config(_media(workspace)))
    out = workspace / "report.csv"
    report_video_candidates.main(["acme", "--format", "csv", "--out", str(out)])

    rows = _read_csv(out)
    states = {(r["language"], r["file"]): r["state"] for r in rows}
    assert states == {
        ("nl", "DrainSticks_NL.mpeg"): "unset",
        ("nl", "Bulbman.mpg"): "confirmed",
        ("nl", "Stray.mpg"): "NOT IN MAPPING",
        ("fr", "DrainSticks_FR.mpeg"): "unset",
    }


def test_a_confirmed_row_carries_the_names_and_an_unset_one_carries_the_hints(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """The two halves of the report's job in one assertion: check what is mapped, decide what is
    not. The winning field on the unset row is a French one holding English, which is the case
    the value/field columns exist for."""
    _patch_client(monkeypatch, _make_config(_media(workspace)))
    out = workspace / "report.csv"
    report_video_candidates.main(["acme", "--format", "csv", "--out", str(out)])

    rows = {r["file"]: r for r in _read_csv(out)}
    assert rows["Bulbman.mpg"]["mapped_product_name.nl"] == "stickylamp"
    assert rows["Bulbman.mpg"]["mapped_logistics_name.fr"] == "Bulb man"

    unset = rows["DrainSticks_NL.mpeg"]
    assert unset["candidate_1_gtin"] == "08713195006796"
    assert unset["candidate_1_value"] == "Drain Sticks 12pc"
    assert unset["candidate_1_field"] == "extras.logistics_name.fr"


def test_top_n_controls_how_many_candidates_each_row_offers(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _patch_client(monkeypatch, _make_config(_media(workspace)))
    out = workspace / "report.csv"
    report_video_candidates.main(["acme", "--format", "csv", "--top-n", "1", "--out", str(out)])

    columns = list(_read_csv(out)[0])
    assert columns[-1] == "candidate_1_field"
    assert "candidate_2_gtin" not in columns


def test_the_workbook_freezes_and_filters_the_header_and_keeps_scores_numeric(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """173 rows is a scrolling, sorting job. A header that scrolls away, or scores stored as text
    that sort 0.9 above 0.83, would make the artefact worse than the CSV it replaces."""
    _patch_client(monkeypatch, _make_config(_media(workspace)))
    out = workspace / "report.xlsx"
    report_video_candidates.main(["acme", "--out", str(out)])

    sheet = openpyxl.load_workbook(out).active
    header = [cell.value for cell in sheet[1]]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == sheet.dimensions
    score = sheet.cell(row=2, column=header.index("candidate_1_score") + 1).value
    gtin = sheet.cell(row=2, column=header.index("candidate_1_gtin") + 1).value
    assert isinstance(score, float)
    assert gtin == "08713195006796", "a GTIN written as a number loses its leading zero"


def test_it_says_what_it_counted(
    workspace: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    _patch_client(monkeypatch, _make_config(_media(workspace)))
    report_video_candidates.main(["acme"])

    err = capsys.readouterr().err
    assert "4 row(s)" in err
    assert "1 confirmed" in err
    assert "2 unset" in err


def test_an_absent_video_library_is_named_rather_than_reported_as_a_broken_mapping(
    workspace: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    """The condition `VideoMapSummary.no_files_found` exists for. Without the line, an operator
    whose folders are on another machine reads a file of NOT ON DISK rows as a mapping to fix."""
    _patch_client(monkeypatch, _make_config(_media(workspace / "elsewhere")))

    assert report_video_candidates.main(["acme"]) == 0
    err = capsys.readouterr().err
    assert "no video files found" in err
    assert "3 NOT ON DISK" in err


# --- Refusals ----------------------------------------------------------------


def test_a_client_with_no_video_mapping_is_a_usage_error(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _patch_client(monkeypatch, _make_config(_media(workspace, map_path=None)))

    assert report_video_candidates.main(["acme"]) == 2


def test_a_client_with_no_media_config_is_a_usage_error(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _patch_client(monkeypatch, _make_config(None))

    assert report_video_candidates.main(["acme"]) == 2


def test_a_top_n_below_one_is_refused_rather_than_writing_a_report_with_no_hints(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _patch_client(monkeypatch, _make_config(_media(workspace)))

    assert report_video_candidates.main(["acme", "--top-n", "0"]) == 2


def test_a_missing_products_file_is_a_read_error_not_an_empty_report(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """An empty candidate pool would render every row with no hints and exit 0 — a report that
    looks finished and answers nothing."""
    _patch_client(monkeypatch, _make_config(_media(workspace)))

    assert report_video_candidates.main(["acme", "--products", "nowhere.json"]) == 1


def test_an_unreadable_mapping_is_a_read_error(
    workspace: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    (workspace / "mapping.yml").write_text("nl: [oops\n", encoding="utf-8")
    _patch_client(monkeypatch, _make_config(_media(workspace)))

    assert report_video_candidates.main(["acme"]) == 1
