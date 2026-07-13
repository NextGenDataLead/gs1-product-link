"""Tests for lib/website_status.py — the create-only eligibility gate.

Fixtures are written with openpyxl (transitional OOXML). The reader is namespace-agnostic
and also handles Strict OOXML, header rows below row 1, and the data sheet sitting beside
other sheets — those irregularities of the real operator export are verified against the
live file during Phase 7 end-to-end checks; here we cover the header auto-detection and
sheet-scan logic with synthetic workbooks.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pytest

from lib.config import WebsiteStatusConfig
from lib.errors import WebsiteStatusError
from lib.website_status import WebsiteStatus, load_website_status

_HEADER = [
    "Artikelnr.",
    "Omschrijving NL",
    "Barcode",
    "Momenteel op Website",
    "Al in Gs1",
    "Link naar site",
]

GTIN13 = "8713195004778"  # as written in the control file (no leading zero)
GTIN14 = "08713195004778"  # canonical key after GTIN-14 normalisation


def _write_xlsx(
    tmp_path: Path,
    rows: list[list[Any]],
    header: list[str] = _HEADER,
    *,
    header_start_row: int = 1,
    extra_sheet_first: bool = False,
) -> str:
    """Write a control-file workbook; optionally offset the header and prepend a sheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if extra_sheet_first:
        sheet.append(["a pivot / summary sheet with no data columns"])
        sheet = workbook.create_sheet("Blad1")
    for offset, name in enumerate(header, start=1):
        sheet.cell(row=header_start_row, column=offset, value=name)
    for r, row in enumerate(rows, start=header_start_row + 1):
        for offset, value in enumerate(row, start=1):
            sheet.cell(row=r, column=offset, value=value)
    path = tmp_path / "website_status.xlsx"
    workbook.save(path)
    return str(path)


def _config(path: str) -> WebsiteStatusConfig:
    return WebsiteStatusConfig(path=path)


def test_blank_and_filled_cells_parse_as_booleans(tmp_path: Path) -> None:
    # Arrange: on-website blank, in-GS1 filled -> eligible.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, None, "*", "https://x/1"]])

    # Act
    statuses = load_website_status(_config(path))

    # Assert (keyed by GTIN-14)
    status = statuses[GTIN14]
    assert status == WebsiteStatus(
        gtin=GTIN14, on_website=False, in_gs1=True, site_link="https://x/1"
    )
    assert status.eligible is True


@pytest.mark.parametrize(
    ("on_website", "in_gs1", "eligible"),
    [
        (None, "*", True),  # not on site, in GS1 -> eligible
        ("x", "*", False),  # already on site -> skip
        (None, None, False),  # not in GS1 -> skip
        ("x", None, False),  # on site and not in GS1 -> skip
    ],
)
def test_eligibility_requires_in_gs1_and_not_on_website(
    tmp_path: Path, on_website: Any, in_gs1: Any, eligible: bool
) -> None:
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, on_website, in_gs1, None]])

    statuses = load_website_status(_config(path))

    assert statuses[GTIN14].eligible is eligible


def test_whitespace_only_cell_counts_as_blank(tmp_path: Path) -> None:
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, "   ", "*", None]])

    status = load_website_status(_config(path))[GTIN14]

    assert status.on_website is False
    assert status.eligible is True


def test_thirteen_digit_barcode_normalised_to_gtin14(tmp_path: Path) -> None:
    # The real control file stores 13-digit barcodes without a leading zero.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, None, "*", None]])

    statuses = load_website_status(_config(path))

    assert GTIN14 in statuses
    assert GTIN13 not in statuses


def test_numeric_barcode_coerced_and_normalised(tmp_path: Path) -> None:
    # openpyxl stores an unquoted barcode as an int; it must key as a GTIN-14 string.
    path = _write_xlsx(tmp_path, [["A1", "Widget", 8713195004778, None, "*", None]])

    statuses = load_website_status(_config(path))

    assert GTIN14 in statuses


def test_header_not_on_first_row_is_auto_detected(tmp_path: Path) -> None:
    # The operator file has a report title/pivot above the table; header starts lower.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, None, "*", None]], header_start_row=4)

    statuses = load_website_status(_config(path))

    assert statuses[GTIN14].eligible is True


def test_data_sheet_found_beside_other_sheets(tmp_path: Path) -> None:
    path = _write_xlsx(
        tmp_path, [["A1", "Widget", GTIN13, None, "*", None]], extra_sheet_first=True
    )

    statuses = load_website_status(_config(path))

    assert statuses[GTIN14].eligible is True


def test_rows_with_blank_barcode_are_skipped(tmp_path: Path) -> None:
    path = _write_xlsx(
        tmp_path,
        [
            ["A1", "Widget", GTIN13, None, "*", None],
            ["A2", "No barcode", None, None, "*", None],
        ],
    )

    statuses = load_website_status(_config(path))

    assert list(statuses) == [GTIN14]


def test_missing_required_column_raises(tmp_path: Path) -> None:
    header = ["Artikelnr.", "Omschrijving NL", "Barcode", "Al in Gs1"]  # no "Momenteel op Website"
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, "*"]], header=header)

    with pytest.raises(WebsiteStatusError, match="required columns"):
        load_website_status(_config(path))


def test_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(WebsiteStatusError, match="cannot read"):
        load_website_status(_config(str(tmp_path / "does_not_exist.xlsx")))


def test_optional_site_link_column_absent(tmp_path: Path) -> None:
    header = ["Artikelnr.", "Omschrijving NL", "Barcode", "Momenteel op Website", "Al in Gs1"]
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13, None, "*"]], header=header)

    status = load_website_status(_config(path))[GTIN14]

    assert status.site_link is None
    assert status.eligible is True
