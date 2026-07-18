"""Unit tests for the GDSN datapool reader (IMPLEMENTATION_SPEC §3 extension)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from lib.gdsn import BuildResult, GdsnSource, build_records, read_workbook
from lib.records import ProductRecord

# A synthesized mini GDSN workbook: 7 header rows, LanguageCode/Value pairs. Unlike the
# real export's earlier fixture, EVERY market row carries BOTH nl and fr — matching
# production, where the market that holds a given value varies by product. That is what
# lets these tests exercise ranked resolution, cross-market inconsistency, and blanks.

_DESC_HEADER = [
    [
        "Gtin",
        "TargetMarketCountryCode",
        "InformationProviderOfTradeItem",
        "TradeItemUnitDescriptorCode",
        "TradeItemDescriptionInformation",
        "TradeItemDescriptionInformation",
        "TradeItemDescriptionInformation",
        "TradeItemDescriptionInformation",
        "TradeItemDescriptionInformation",
    ],
    [
        None,
        None,
        None,
        None,
        "DescriptionShort[0]",
        "DescriptionShort[0]",
        "DescriptionShort[1]",
        "DescriptionShort[1]",
        "BrandNameInformation",
    ],
    [None, None, None, None, "LanguageCode", "Value", "LanguageCode", "Value", "BrandName"],
    [None] * 9,
    [None] * 9,
    [None] * 9,
    [
        "GTIN (3059)",
        "Country (3179)",
        "Provider (3088)",
        "Unit (3074)",
        "Short product name (3297)",
        "Short product name (3297)",
        "Short product name (3297)",
        "Short product name (3297)",
        "Brand Name (3336)",
    ],
]


def _drow(gtin: str, market: str, *tail: object) -> list[object]:
    """Build a data row with the shared GLN + consumer-unit key columns."""
    return [gtin, market, "GLN", "BASE_UNIT_OR_EACH", *tail]


_DESC_DATA = [
    # 08713195007359: both markets carry nl AND fr. nl agrees; fr DIFFERS between markets
    # (528 "Support 528" vs 056 "Support 056") — the cross-market inconsistency case. With
    # priority [528, 056], fr resolves to the 528 value.
    _drow("08713195007359", "528", "nl", "Rugsteun NL", "fr", "Support 528", "Noviplast"),
    _drow("08713195007359", "056", "nl", "Rugsteun NL", "fr", "Support 056", "Noviplast"),
    # 08713195000794: nl in both markets, fr in neither — the value_blank case for fr.
    _drow("08713195000794", "528", "nl", "Alleen NL", None, None, "Noviplast"),
    _drow("08713195000794", "056", "nl", "Alleen NL", None, None, "Noviplast"),
    # 09999999999999: fr only, no nl in any market — E5 (missing default-language name).
    _drow("09999999999999", "056", "fr", "Solo FR", None, None, "Noviplast"),
    [None] * 9,  # E4: empty row
]

_MEAS_HEADER = [
    [
        "Gtin",
        "TargetMarketCountryCode",
        "InformationProviderOfTradeItem",
        "TradeItemUnitDescriptorCode",
        "TradeItemMeasurements",
        "TradeItemMeasurements",
    ],
    [None, None, None, None, "NetContent[0]", "NetContent[0]"],
    [None, None, None, None, "MeasurementUnitCode", "Value"],
    [None] * 6,
    [None] * 6,
    [None] * 6,
    [
        "GTIN (3059)",
        "Country (3179)",
        "Provider (3088)",
        "Unit (3074)",
        "Net Content (3510)",
        "Net Content (3510)",
    ],
]
_MEAS_DATA = [
    _drow("08713195007359", "528", "H87", "4"),
]


def _write_workbook(tmp_path: Path) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    desc = wb.create_sheet("TradeItemDescription")
    for row in [*_DESC_HEADER, *_DESC_DATA]:
        desc.append(row)
    meas = wb.create_sheet("TradeItemMeasurements")
    for row in [*_MEAS_HEADER, *_MEAS_DATA]:
        meas.append(row)
    path = tmp_path / "gdsn.xlsx"
    wb.save(path)
    return str(path)


_GDSN_MAP = {
    "product_name": GdsnSource(sheet="TradeItemDescription", attribute="3297", localised=True),
    "brand": GdsnSource(sheet="TradeItemDescription", attribute="3336"),
    "net_content": GdsnSource(sheet="TradeItemMeasurements", attribute="3510", with_unit=True),
}
_MARKET_PRIORITY = ["528", "056"]
_LANGUAGES = ["nl", "fr"]


def test_header_detection_and_column_parsing(tmp_path: Path) -> None:
    sheets = read_workbook(_write_workbook(tmp_path))

    desc = sheets["TradeItemDescription"]
    value_col = next(c for c in desc.columns if c.index == 5)
    assert value_col.leaf_name == "Value"
    assert value_col.attr_id == "3297"
    assert value_col.group_path == ("TradeItemDescriptionInformation", "DescriptionShort[0]")
    brand_col = next(c for c in desc.columns if c.index == 8)
    assert brand_col.leaf_name == "BrandName"
    assert brand_col.matches_attribute("3336")


def test_pickers_resolve_language_and_unit(tmp_path: Path) -> None:
    sheets = read_workbook(_write_workbook(tmp_path))
    desc = sheets["TradeItemDescription"]
    meas = sheets["TradeItemMeasurements"]

    assert desc.pick_localised("08713195007359", "528", "3297", "nl") == "Rugsteun NL"
    # The 528 row now carries fr too — the real export's shape. The old fixture asserted
    # this was None, which is exactly the falsehood that let the 1:1 map look correct.
    assert desc.pick_localised("08713195007359", "528", "3297", "fr") == "Support 528"
    assert desc.pick_localised("08713195007359", "056", "3297", "fr") == "Support 056"
    assert desc.pick_scalar("08713195007359", "528", "3336") == "Noviplast"
    assert meas.pick_scalar("08713195007359", "528", "3510", with_unit=True) == "4 H87"


def test_build_records_joins_markets_into_one_record(tmp_path: Path) -> None:
    # The same GTIN across markets aggregates into ONE record. nl is identical in both
    # markets; fr differs, and ranked resolution takes the highest-priority market (528).
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, _MARKET_PRIORITY, _LANGUAGES, "nl")

    good = [r for r in result.records if r.gtin == "08713195007359"]
    assert len(good) == 1
    assert good[0].product_name.values == {"nl": "Rugsteun NL", "fr": "Support 528"}
    assert good[0].brand == "Noviplast"
    assert good[0].net_content == "4 H87"


def test_build_records_reports_cross_market_inconsistency(tmp_path: Path) -> None:
    # fr differs between 528 ("Support 528") and 056 ("Support 056"): the tool takes the
    # ranked winner and reports the disagreement rather than silently choosing.
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, _MARKET_PRIORITY, _LANGUAGES, "nl")

    issues = [i for i in result.issues if i.issue == "value_inconsistent_across_markets"]
    assert len(issues) == 1
    assert issues[0].field == "product_name.fr"
    assert issues[0].gtin == "08713195007359"
    assert issues[0].value == "Support 528"  # the value actually used
    assert "528=" in issues[0].detail and "056=" in issues[0].detail
    # nl agrees across markets, so it is NOT reported.
    assert not any(i.field == "product_name.nl" for i in issues)


def test_inconsistency_ignores_case_and_whitespace_only_differences(tmp_path: Path) -> None:
    # "Rugsteun"/"rugsteun" is not a content disagreement, and the title CSS uppercases
    # anyway. Reporting it would bury the substantive conflicts in noise.
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    desc = wb.create_sheet("TradeItemDescription")
    for row in _DESC_HEADER:
        desc.append(row)
    desc.append(_drow("08713195007359", "528", "nl", "Rugsteun", None, None, "Noviplast"))
    desc.append(_drow("08713195007359", "056", "nl", "  rugsteun ", None, None, "Noviplast"))
    path = tmp_path / "case.xlsx"
    wb.save(path)

    result = build_records(read_workbook(str(path)), _GDSN_MAP, _MARKET_PRIORITY, ["nl"], "nl")

    assert not any(i.issue == "value_inconsistent_across_markets" for i in result.issues)


def test_build_records_reports_blank_for_published_field(tmp_path: Path) -> None:
    # 08713195000794 has nl in both markets but fr in neither: product_name.fr is blank
    # everywhere, so it is reported (report_issues defaults True for a published field).
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, _MARKET_PRIORITY, _LANGUAGES, "nl")

    blanks = [i for i in result.issues if i.issue == "value_blank"]
    assert any(i.gtin == "08713195000794" and i.field == "product_name.fr" for i in blanks)
    # The product still builds with its nl name — a blank is a report, not a failure.
    rec = next(r for r in result.records if r.gtin == "08713195000794")
    assert rec.product_name.values == {"nl": "Alleen NL"}


def test_report_issues_false_suppresses_both_blank_and_inconsistency(tmp_path: Path) -> None:
    # A generator-input field (report_issues=False) is silent for BOTH kinds — its gaps and
    # cross-market conflicts are the generator's future work, not today's source-fix queue.
    # The fixture GTIN 08713195007359 has an fr inconsistency and 000794 has an fr blank; a
    # quiet product_name must surface neither.
    sheets = read_workbook(_write_workbook(tmp_path))
    quiet_map = {
        **_GDSN_MAP,
        "product_name": GdsnSource(
            sheet="TradeItemDescription", attribute="3297", localised=True, report_issues=False
        ),
    }

    result = build_records(sheets, quiet_map, _MARKET_PRIORITY, _LANGUAGES, "nl")

    product_name_issues = [i for i in result.issues if i.field.startswith("product_name")]
    assert product_name_issues == []
    # The value is still resolved and the record still built — only the reporting is muted.
    rec = next(r for r in result.records if r.gtin == "08713195007359")
    assert rec.product_name.values["fr"] == "Support 528"


def test_build_records_higher_priority_market_wins(tmp_path: Path) -> None:
    # Reversing the priority flips which market supplies fr — proving the order decides.
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, ["056", "528"], _LANGUAGES, "nl")

    rec = next(r for r in result.records if r.gtin == "08713195007359")
    assert rec.product_name.values["fr"] == "Support 056"


def _prefix_map() -> dict[str, GdsnSource]:
    return {
        **_GDSN_MAP,
        "product_name": GdsnSource(
            sheet="TradeItemDescription",
            attribute="3297",
            localised=True,
            strip_prefix="Noviplast ",
        ),
    }


def _build_named(tmp_path: Path, nl: str, fr: str) -> BuildResult:
    """Build a one-GTIN workbook whose product name is ``nl``/``fr``, and parse it."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    desc = wb.create_sheet("TradeItemDescription")
    for row in _DESC_HEADER:
        desc.append(row)
    desc.append(_drow("08713195007359", "528", "nl", nl, None, None, "Noviplast"))
    desc.append(_drow("08713195007359", "056", "fr", fr, None, None, "Noviplast"))
    path = tmp_path / "prefix.xlsx"
    wb.save(path)
    return build_records(
        read_workbook(str(path)), _prefix_map(), _MARKET_PRIORITY, _LANGUAGES, "nl"
    )


def _named(tmp_path: Path, nl: str, fr: str) -> ProductRecord:
    return _build_named(tmp_path, nl, fr).records[0]


def test_strip_prefix_removes_the_brand_from_the_name(tmp_path: Path) -> None:
    rec = _named(tmp_path, "Noviplast Microvezeldoek stof", "Noviplast Super5 microfibre")

    assert rec.product_name.values == {"nl": "Microvezeldoek stof", "fr": "Super5 microfibre"}
    assert rec.brand == "Noviplast"  # brand is its own field, unaffected


def test_strip_prefix_leaves_genuinely_unprefixed_names_alone(
    tmp_path: Path, caplog: pytest.LogCaptureFixture
) -> None:
    """Not every product name repeats the brand — those must pass through silently."""
    with caplog.at_level("WARNING", logger="lib.gdsn"):
        rec = _named(tmp_path, "Super Glove", "Garden Clipper")

    assert rec.product_name.values == {"nl": "Super Glove", "fr": "Garden Clipper"}
    assert caplog.text == ""  # no false-positive typo report


@pytest.mark.parametrize(
    "misspelt",
    [
        "Noviplat Snoeischaar metaal grijs",  # dropped 's' — real, in the pilot export
        "Nociplast Bouteilles à marinade",  # v -> c
        "Novilplast Détecteur de mouvement",  # inserted 'l'
        "NoviplastSnijplanken kunststof grijs",  # missing space
    ],
)
def test_strip_prefix_reports_misspellings_but_never_repairs_them(
    tmp_path: Path, caplog: pytest.LogCaptureFixture, misspelt: str
) -> None:
    """A near-miss is a defect in the source datapool, not something to silently fix.

    All four shapes are real values from the pilot export. Repairing them here would hide
    the defect while the wrong text stays authoritative upstream, so the value passes
    through unchanged and the operator is told to fix it at source.
    """
    with caplog.at_level("WARNING", logger="lib.gdsn"):
        rec = _named(tmp_path, misspelt, "Noviplast Bon nom")

    assert rec.product_name.values["nl"] == misspelt  # unchanged, not "corrected"
    assert rec.product_name.values["fr"] == "Bon nom"  # the well-formed one still strips
    assert "resembles but does not match" in caplog.text
    assert "08713195007359" in caplog.text  # the GTIN, so it can be found in MyGS1


def _length_map(limit: int) -> dict[str, GdsnSource]:
    return {
        **_GDSN_MAP,
        "product_name": GdsnSource(
            sheet="TradeItemDescription", attribute="3297", localised=True, max_length=limit
        ),
    }


def _build_with(tmp_path: Path, gdsn_map: dict[str, GdsnSource], nl: str) -> BuildResult:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    desc = wb.create_sheet("TradeItemDescription")
    for row in _DESC_HEADER:
        desc.append(row)
    desc.append(_drow("08713195007359", "528", "nl", nl, None, None, "Noviplast"))
    path = tmp_path / "len.xlsx"
    wb.save(path)
    # nl only: this helper builds a single-language row, so declaring only nl keeps the
    # length/prefix checks isolated from fr value_blank noise they are not about.
    return build_records(read_workbook(str(path)), gdsn_map, _MARKET_PRIORITY, ["nl"], "nl")


def test_max_length_reports_but_keeps_the_value(tmp_path: Path) -> None:
    """Over-long is reported and kept verbatim — truncating would sever a sentence mid-word
    on the page while the value stays too long in MyGS1 and returns on the next export."""
    long_value = "x" * 200

    result = _build_with(tmp_path, _length_map(120), long_value)

    assert result.records[0].product_name.values["nl"] == long_value  # untruncated
    issues = [i for i in result.issues if i.issue == "value_too_long"]
    assert len(issues) == 1
    assert issues[0].field == "product_name.nl"
    # Named as the source system names it — "product_name" exists only in this codebase.
    assert issues[0].source == "TradeItemDescription attr 3297"
    assert issues[0].gtin == "08713195007359"
    assert "200 characters, longer than the 120 expected" in issues[0].detail
    assert result.errors == []  # non-fatal


def test_max_length_silent_at_or_under_the_limit(tmp_path: Path) -> None:
    result = _build_with(tmp_path, _length_map(120), "x" * 120)

    assert [i for i in result.issues if i.issue == "value_too_long"] == []


def test_no_max_length_means_no_expectation(tmp_path: Path) -> None:
    result = _build_with(tmp_path, _GDSN_MAP, "x" * 5000)

    assert result.issues == []


def test_length_is_measured_after_the_prefix_is_stripped(tmp_path: Path) -> None:
    """The brand prefix is removed before rendering, so it must not count toward the slot."""
    gdsn_map = {
        **_GDSN_MAP,
        "product_name": GdsnSource(
            sheet="TradeItemDescription",
            attribute="3297",
            localised=True,
            strip_prefix="Noviplast ",
            max_length=20,
        ),
    }

    result = _build_with(tmp_path, gdsn_map, "Noviplast Rugsteun")  # 18 once stripped

    assert result.records[0].product_name.values["nl"] == "Rugsteun"
    assert [i for i in result.issues if i.issue == "value_too_long"] == []


def test_strip_prefix_report_reaches_the_result_warnings(tmp_path: Path) -> None:
    """The note must land in BuildResult.warnings, not only in the log.

    parse_export's summary counts result.warnings; a note that only logs makes the run
    print four warnings and then report "0 warnings" — which is how a warning gets ignored.
    """
    result = _build_named(tmp_path, "Noviplat Snoeischaar", "Noviplast Bon nom")

    # (The fixture omits the measurements sheet, so an unrelated "absent source" warning
    # rides along — filter to the one under test.)
    notes = [w for w in result.warnings if "resembles but does not match" in w]
    assert len(notes) == 1
    assert "08713195007359" in notes[0]
    assert result.errors == []  # non-fatal: the record is still built
    assert result.records


def test_build_records_missing_default_language_is_error(tmp_path: Path) -> None:
    # E5: GTIN 09999999999999 has fr in market 056 but no nl in any market.
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, _MARKET_PRIORITY, _LANGUAGES, "nl")

    assert not any(r.gtin == "09999999999999" for r in result.records)
    assert any("09999999999999" in e and "product_name.nl" in e for e in result.errors)


def test_build_records_missing_required_source_errors(tmp_path: Path) -> None:
    # E17 (required): product_name mapped to a sheet that isn't present.
    sheets = read_workbook(_write_workbook(tmp_path))
    bad_map = {
        "product_name": GdsnSource(sheet="Nope", attribute="3297", localised=True),
        "brand": GdsnSource(sheet="TradeItemDescription", attribute="3336"),
    }

    result = build_records(sheets, bad_map, _MARKET_PRIORITY, _LANGUAGES, "nl")

    assert result.records == []
    assert any("product_name" in e and "Nope" in e for e in result.errors)


def test_build_records_empty_market_priority_is_error(tmp_path: Path) -> None:
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, [], _LANGUAGES, "nl")

    assert result.records == []
    assert any("market_priority" in e for e in result.errors)


def test_build_records_default_language_not_in_languages_is_error(tmp_path: Path) -> None:
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, _MARKET_PRIORITY, ["fr"], "nl")

    assert result.records == []
    assert any("default_language" in e for e in result.errors)


def test_reference_sheets_without_data_are_skipped(tmp_path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ref = wb.create_sheet("Access Control Group")
    ref.append(["This sheet is reference only.", "", "", ""])
    ref.append(["Name", "x", "y", "z"])
    desc = wb.create_sheet("TradeItemDescription")
    for row in [*_DESC_HEADER, *_DESC_DATA]:
        desc.append(row)
    path = tmp_path / "ref.xlsx"
    wb.save(path)

    sheets = read_workbook(str(path))

    assert "Access Control Group" not in sheets
    assert "TradeItemDescription" in sheets


def test_scalars_come_from_the_highest_priority_market_with_a_row(tmp_path: Path) -> None:
    # Scalars (brand, net_content) are picked by walking market_priority, not by a
    # default-language market. net_content lives only in market 528's measurements row, so
    # it resolves regardless of which language is default.
    sheets = read_workbook(_write_workbook(tmp_path))

    result = build_records(sheets, _GDSN_MAP, _MARKET_PRIORITY, _LANGUAGES, "fr")

    record = next(r for r in result.records if r.gtin == "08713195007359")
    assert record.brand == "Noviplast"
    assert record.net_content == "4 H87"  # from market 528, though fr is default


def _write_generator_inputs_workbook(tmp_path: Path) -> str:
    """A workbook exercising the three content-generator input shapes as ``gdsn_extras``:
    a localised variation (3332), a with-unit dimension (3498), and a segment-matched
    scalar material (no numeric attr id — matched by the ``Material`` path segment)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    desc = wb.create_sheet("TradeItemDescription")
    for row in [
        # name (3301) + brand (3336) + variation (3332), name & variation as LanguageCode/Value
        ["Gtin", "TargetMarketCountryCode", "InformationProviderOfTradeItem",
         "TradeItemUnitDescriptorCode", "Info", "Info", "BrandNameInformation", "Info", "Info"],
        [None, None, None, None, "Name[0]", "Name[0]", None, "Variation[0]", "Variation[0]"],
        [None, None, None, None, "LanguageCode", "Value", "BrandName", "LanguageCode", "Value"],
        ["GTIN (3059)", "Country (3179)", "Provider (3088)", "Unit (3074)",
         "Functional (3301)", "Functional (3301)", "Brand (3336)",
         "Variation (3332)", "Variation (3332)"],
        _drow("08713195000794", "528", "nl", "Voegstrijker", "Noviplast", "nl", "Set"),
    ]:
        desc.append(row)

    meas = wb.create_sheet("TradeItemMeasurements")
    for row in [
        ["Gtin", "TargetMarketCountryCode", "InformationProviderOfTradeItem",
         "TradeItemUnitDescriptorCode", "TradeItemMeasurements", "TradeItemMeasurements"],
        [None, None, None, None, "Height[0]", "Height[0]"],
        [None, None, None, None, "MeasurementUnitCode", "Value"],
        ["GTIN (3059)", "Country (3179)", "Provider (3088)", "Unit (3074)",
         "Height (3498)", "Height (3498)"],
        _drow("08713195000794", "528", "MMT", "250"),
    ]:
        meas.append(row)

    brick = wb.create_sheet("BrickGPCCommercialData")
    for row in [
        # Material is Information[0]/Material[0]/Value with a non-numeric "(4.012)" label,
        # so it carries no attr_id and must be matched by the "Material" path segment.
        ["Gtin", "TargetMarketCountryCode", "InformationProviderOfTradeItem",
         "TradeItemUnitDescriptorCode", "Information[0]"],
        [None, None, None, None, "Material[0]"],
        [None, None, None, None, "Value"],
        ["GTIN (3059)", "Country (3179)", "Provider (3088)", "Unit (3074)", "Material (4.012)"],
        _drow("08713195000794", "528", "kunststof"),
    ]:
        brick.append(row)

    path = tmp_path / "generator_inputs.xlsx"
    wb.save(path)
    return str(path)


def test_gdsn_extras_carry_generator_inputs(tmp_path: Path) -> None:
    # The generator's parser inputs ride in extras, one entry per shape: a localised token
    # collapsed to the default language, a dimension with its unit code preserved for later
    # decoding, and a segment-matched scalar. None of these are published fields.
    sheets = read_workbook(_write_generator_inputs_workbook(tmp_path))
    gdsn_map = {
        "product_name": GdsnSource(sheet="TradeItemDescription", attribute="3301", localised=True),
        "brand": GdsnSource(sheet="TradeItemDescription", attribute="3336"),
    }
    gdsn_extras = {
        "product_variation": GdsnSource(
            sheet="TradeItemDescription", attribute="3332", localised=True
        ),
        "dim_height": GdsnSource(sheet="TradeItemMeasurements", attribute="3498", with_unit=True),
        "material": GdsnSource(sheet="BrickGPCCommercialData", attribute="Material"),
    }

    result = build_records(
        sheets, gdsn_map, ["528"], ["nl"], "nl", gdsn_extras=gdsn_extras
    )

    record = next(r for r in result.records if r.gtin == "08713195000794")
    assert record.extras["product_variation"] == "Set"
    assert record.extras["dim_height"] == "250 MMT"  # unit code kept for lib/units decoding
    assert record.extras["material"] == "kunststof"


def _write_multivalue_workbook(tmp_path: Path) -> str:
    """A workbook whose 1067 attribute spreads USPs across two repeated slots."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    desc = wb.create_sheet("TradeItemDescription")
    for row in [
        ["Gtin", "TargetMarketCountryCode", "InformationProviderOfTradeItem",
         "TradeItemUnitDescriptorCode", "Info", "Info", "BrandNameInformation"],
        [None, None, None, None, "Name[0]", "Name[0]", None],
        [None, None, None, None, "LanguageCode", "Value", "BrandName"],
        ["GTIN (3059)", "Country (3179)", "Provider (3088)", "Unit (3074)",
         "Functional (3301)", "Functional (3301)", "Brand (3336)"],
        _drow("08713195007717", "528", "nl", "Hogedrukreiniger", "Noviplast"),
    ]:
        desc.append(row)

    mi = wb.create_sheet("MarketingInformation")
    for row in [
        ["Gtin", "TargetMarketCountryCode", "InformationProviderOfTradeItem",
         "TradeItemUnitDescriptorCode", "MI", "MI", "MI", "MI"],
        [None, None, None, None, "FeatureBenefit[0]", "FeatureBenefit[0]",
         "FeatureBenefit[1]", "FeatureBenefit[1]"],
        [None, None, None, None, "LanguageCode", "Value", "LanguageCode", "Value"],
        ["GTIN (3059)", "Country (3179)", "Provider (3088)", "Unit (3074)",
         "Feature (1067)", "Feature (1067)", "Feature (1067)", "Feature (1067)"],
        _drow("08713195007717", "528", "nl", "Eerste USP", "nl", "Tweede USP"),
    ]:
        mi.append(row)

    path = tmp_path / "multivalue.xlsx"
    wb.save(path)
    return str(path)


def test_multivalue_joins_all_slots(tmp_path: Path) -> None:
    # 1067 spreads USPs across TradeItemFeatureBenefit[0] and [1]; multivalue joins both with a
    # newline so the generator can split them into one ranked list (single-value picking would
    # silently drop the second USP).
    sheets = read_workbook(_write_multivalue_workbook(tmp_path))
    gdsn_map = {
        "product_name": GdsnSource(sheet="TradeItemDescription", attribute="3301", localised=True),
        "brand": GdsnSource(sheet="TradeItemDescription", attribute="3336"),
        "description_long": GdsnSource(
            sheet="MarketingInformation", attribute="1067", localised=True, multivalue=True
        ),
    }

    result = build_records(sheets, gdsn_map, ["528"], ["nl"], "nl")

    record = next(r for r in result.records if r.gtin == "08713195007717")
    assert record.description_long is not None
    assert record.description_long.get("nl") == "Eerste USP\nTweede USP"
