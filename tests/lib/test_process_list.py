"""Tests for lib/process_list.py — the operator's list of GTINs to process.

Most fixtures are written with openpyxl (transitional OOXML), which covers header
auto-detection and the sheet scan. Strict OOXML — the format the real operator files are
saved in, and the one openpyxl reads zero sheets from — is covered by a hand-built workbook
at the bottom of this file, because it is the irregularity that actually bit and nothing
else in CI exercises it.

The behaviour under test is deliberately narrow: **every GTIN in the file is processed.**
No cell values are interpreted, so the tests below assert that extra columns are ignored
regardless of what they contain — including the words the previous status-column reader
silently got backwards.
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from lib.config import ProcessListConfig
from lib.errors import ProcessListError
from lib.process_list import ProcessListSheet, load_process_list, read_process_list

_HEADER = ["Artikelnr.", "Omschrijving NL", "Barcode"]

GTIN13 = "8713195004778"  # as written in the file (no leading zero)
GTIN14 = "08713195004778"  # canonical key after GTIN-14 normalisation


def _write_xlsx(
    tmp_path: Path,
    rows: list[list[Any]],
    header: list[str] = _HEADER,
    *,
    header_start_row: int = 1,
    extra_sheet_first: bool = False,
) -> str:
    """Write a process-list workbook; optionally offset the header and prepend a sheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if extra_sheet_first:
        sheet.append(["a pivot / summary sheet with no data columns"])
        sheet = workbook.create_sheet("Blad1")
    for offset, name in enumerate(header, start=1):
        sheet.cell(row=header_start_row, column=offset, value=name)
    for r, row in enumerate(rows, start=header_start_row + 1):
        for offset, value in enumerate(row, start=1):
            sheet.cell(row=r, column=offset, value=value)
    path = tmp_path / "process-list.xlsx"
    workbook.save(path)
    return str(path)


def _config(path: str, gtin_column: str = "Barcode") -> ProcessListConfig:
    return ProcessListConfig(path=path, gtin_column=gtin_column)


def test_every_listed_gtin_is_returned(tmp_path: Path) -> None:
    # Arrange
    path = _write_xlsx(
        tmp_path,
        [["A1", "Widget", GTIN13], ["A2", "Gadget", "8713195004779"]],
    )

    # Act
    listed = load_process_list(_config(path))

    # Assert
    assert listed == frozenset({GTIN14, "08713195004779"})


@pytest.mark.parametrize("marker", ["x", "X", "no", "nee", "FALSE", "0", "", None])
def test_extra_columns_are_ignored_whatever_they_contain(tmp_path: Path, marker: Any) -> None:
    """Membership is the whole rule — a status-looking column must not change the outcome.

    Regression: the previous reader treated any non-blank cell as True, so a file saying
    ``no`` meant the opposite of the word, silently and in both directions.
    """
    # Arrange
    path = _write_xlsx(
        tmp_path,
        [["A1", "Widget", GTIN13, marker]],
        header=[*_HEADER, "Op website"],
    )

    # Act
    listed = load_process_list(_config(path))

    # Assert
    assert listed == frozenset({GTIN14})


def test_thirteen_digit_barcode_normalised_to_gtin14(tmp_path: Path) -> None:
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13]])

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})


def test_numeric_barcode_coerced_and_normalised(tmp_path: Path) -> None:
    """A barcode cell typed as a number must not arrive as ``8713195004778.0``."""
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", int(GTIN13)]])

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})


def test_duplicate_rows_collapse(tmp_path: Path) -> None:
    # Arrange: the same product twice, once 13-digit and once 14-digit.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13], ["A1", "Widget", GTIN14]])

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})


def test_header_not_on_first_row_is_auto_detected(tmp_path: Path) -> None:
    # Arrange: a report title sits above the table, as in the real export.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13]], header_start_row=4)

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})


def test_data_sheet_found_beside_other_sheets(tmp_path: Path) -> None:
    # Arrange: a pivot/summary sheet precedes the data sheet.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13]], extra_sheet_first=True)

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})


def test_rows_with_blank_barcode_are_skipped(tmp_path: Path) -> None:
    # Arrange: trailing total/blank rows carry no barcode.
    path = _write_xlsx(
        tmp_path,
        [["A1", "Widget", GTIN13], ["", "Totaal", None], ["", "", "   "]],
    )

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})


def test_custom_gtin_column_name(tmp_path: Path) -> None:
    """A client may label the column whatever they like — only the name is configured."""
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13]], header=["Artikelnr.", "Naam", "EAN"])

    # Act / Assert
    assert load_process_list(_config(path, gtin_column="EAN")) == frozenset({GTIN14})


def test_missing_gtin_column_raises(tmp_path: Path) -> None:
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget"]], header=["Artikelnr.", "Omschrijving NL"])

    # Act / Assert
    with pytest.raises(ProcessListError, match="no sheet with a 'Barcode' column"):
        load_process_list(_config(path))


def test_column_present_but_no_gtins_raises(tmp_path: Path) -> None:
    """An empty list is an error, not an empty run — it would publish nothing silently."""
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", None]])

    # Act / Assert
    with pytest.raises(ProcessListError, match="no GTINs under it"):
        load_process_list(_config(path))


def test_missing_file_raises(tmp_path: Path) -> None:
    # Act / Assert
    with pytest.raises(ProcessListError, match="cannot read process list"):
        load_process_list(_config(str(tmp_path / "does_not_exist.xlsx")))


# --- The grid: what the editing surface reads ---------------------------------
#
# ``read_process_list`` is the same read, stopping one step earlier — before the GTIN column is
# picked out. The shell used to do this read itself, with openpyxl and a header fixed at row 1,
# which meant a real Strict-OOXML file with a title row above the table loaded in a run and
# failed on screen. One reader, so they cannot disagree.


def test_the_grid_finds_a_header_below_row_one(tmp_path: Path) -> None:
    """The bug. A report title above the table is the shape of the real operator file."""
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13]], header_start_row=4)

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.header == _HEADER
    assert sheet.rows == [["A1", "Widget", GTIN13]]
    assert sheet.gtin_index == 2


def test_the_grid_skips_a_sheet_without_the_column(tmp_path: Path) -> None:
    # Arrange: a pivot/summary sheet precedes the data sheet.
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13]], extra_sheet_first=True)

    # Act / Assert
    assert read_process_list(_config(path)).rows == [["A1", "Widget", GTIN13]]


def test_the_grid_keeps_every_column_in_spreadsheet_order_past_column_z(tmp_path: Path) -> None:
    """Thirty columns, so ``AA`` is real.

    Sorting column letters as strings puts ``AA`` (column 27) between ``A`` and ``B``. The grid
    would come back reordered and ``gtin_index`` would point at somebody else's column — a save
    then writes the barcodes into it.
    """
    # Arrange
    header = ["Artikelnr.", "Omschrijving NL", "Barcode", *(f"note {n}" for n in range(4, 31))]
    row = ["A1", "Widget", GTIN13, *(f"v{n}" for n in range(4, 31))]
    path = _write_xlsx(tmp_path, [row], header=header)

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.header == header
    assert sheet.rows == [row]
    assert sheet.gtin_index == 2


def test_a_column_with_a_blank_header_cell_is_still_carried(tmp_path: Path) -> None:
    """A save rewrites this grid, so a column dropped here is deleted from the operator's file."""
    # Arrange: column D has data but no header — the grid takes the union of both.
    path = _write_xlsx(
        tmp_path,
        [["A1", "Widget", GTIN13, "a note nobody labelled"]],
        header=_HEADER,
    )

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.header == [*_HEADER, ""]
    assert sheet.rows == [["A1", "Widget", GTIN13, "a note nobody labelled"]]


def test_a_fully_blank_row_is_dropped(tmp_path: Path) -> None:
    # Arrange: a spacer row and a whitespace-only row, as spreadsheets accumulate.
    path = _write_xlsx(
        tmp_path,
        [["A1", "Widget", GTIN13], [None, None, None], ["  ", "  ", "  "], ["A2", "Gadget", "871"]],
    )

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.rows == [["A1", "Widget", GTIN13], ["A2", "Gadget", "871"]]


def test_the_grids_gtins_are_the_ones_the_run_reads(tmp_path: Path) -> None:
    """The whole point of one reader: the screen and the pipeline agree about the same file."""
    # Arrange
    path = _write_xlsx(
        tmp_path,
        [["A1", "Widget", GTIN13], ["A2", "Gadget", "8713195004779"], ["", "Totaal", None]],
        header_start_row=3,
    )
    config = _config(path)

    # Act / Assert
    assert read_process_list(config).listed_gtins() == load_process_list(config)


def test_the_grid_does_not_refuse_an_empty_list(tmp_path: Path) -> None:
    """An empty grid is displayable and fixable; refusing it belongs to the callers that act.

    ``load_process_list`` refuses because it would plan nothing; ``save_sheet`` refuses because
    it would write it. Refusing to *show* it just leaves the operator with an error and no way
    to see what is wrong.
    """
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", None]])

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.listed_gtins() == frozenset()
    with pytest.raises(ProcessListError, match="no GTINs under it"):
        load_process_list(_config(path))


def test_gtin14_at_reads_the_configured_column(tmp_path: Path) -> None:
    # Arrange
    path = _write_xlsx(tmp_path, [["A1", "Widget", GTIN13], ["A2", "Gadget", None]])

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.gtin14_at(0) == GTIN14
    assert sheet.gtin14_at(1) is None


def test_keeping_and_without_survive_the_move(tmp_path: Path) -> None:
    """``ProcessListSheet`` moved here from ui/; the list arithmetic is the same arithmetic."""
    # Arrange
    sheet = read_process_list(
        _config(_write_xlsx(tmp_path, [["A1", "a", GTIN13], ["A2", "b", "8713195004779"]]))
    )

    # Act / Assert
    assert isinstance(sheet, ProcessListSheet)
    assert sheet.keeping({0}).rows == [["A1", "a", GTIN13]]
    assert sheet.without({0}).rows == [["A2", "b", "8713195004779"]]
    assert len(sheet.rows) == 2, "the original is untouched"


# --- Strict Open XML ----------------------------------------------------------
#
# The format the real operator files are saved in. openpyxl reads **zero sheets** from one, so
# nothing written with openpyxl can cover it; this workbook is built by hand. Everything differs
# only by namespace URI, which is exactly why the reader matches on local tag names.

_STRICT_MAIN = "http://purl.oclc.org/ooxml/spreadsheetml/main"
_STRICT_REL = "http://purl.oclc.org/ooxml/officeDocument/relationships"
_PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"
_OFFICE_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_SPREADSHEET = "application/vnd.openxmlformats-officedocument.spreadsheetml"


def _write_strict_xlsx(tmp_path: Path, strings: list[str], rows: list[list[int | None]]) -> str:
    """Write a Strict-OOXML workbook whose cells index ``strings``; ``None`` leaves a cell out.

    A complete package, content types and root relationships included, so that openpyxl gets far
    enough to fail the way it fails on a real one — reading the workbook and finding no sheets in
    it — rather than tripping over a missing part this fixture never wrote.
    """

    def cell(column: int, row: int, index: int) -> str:
        letters = ""
        n = column
        while n:
            n, remainder = divmod(n - 1, 26)
            letters = chr(ord("A") + remainder) + letters
        return f'<c r="{letters}{row}" t="s"><v>{index}</v></c>'

    body = "".join(
        f'<row r="{r}">'
        + "".join(cell(c, r, i) for c, i in enumerate(values, start=1) if i is not None)
        + "</row>"
        for r, values in enumerate(rows, start=1)
    )
    path = tmp_path / "strict.xlsx"
    with zipfile.ZipFile(path, "w") as zf:
        zf.writestr(
            "[Content_Types].xml",
            f'<Types xmlns="{_CONTENT_TYPES}">'
            '<Default Extension="rels" '
            'ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            f'<Override PartName="/xl/workbook.xml" ContentType="{_SPREADSHEET}.sheet.main+xml"/>'
            f'<Override PartName="/xl/worksheets/sheet1.xml" '
            f'ContentType="{_SPREADSHEET}.worksheet+xml"/>'
            f'<Override PartName="/xl/sharedStrings.xml" '
            f'ContentType="{_SPREADSHEET}.sharedStrings+xml"/></Types>',
        )
        zf.writestr(
            "_rels/.rels",
            f'<Relationships xmlns="{_PACKAGE_REL}"><Relationship Id="rId1" '
            f'Type="{_OFFICE_REL}/officeDocument" Target="xl/workbook.xml"/></Relationships>',
        )
        zf.writestr(
            "xl/workbook.xml",
            f'<workbook xmlns="{_STRICT_MAIN}" xmlns:r="{_STRICT_REL}">'
            f'<sheets><sheet name="Blad1" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        zf.writestr(
            "xl/_rels/workbook.xml.rels",
            f'<Relationships xmlns="{_PACKAGE_REL}"><Relationship Id="rId1" '
            f'Type="{_OFFICE_REL}/worksheet" Target="worksheets/sheet1.xml"/></Relationships>',
        )
        zf.writestr(
            "xl/sharedStrings.xml",
            f'<sst xmlns="{_STRICT_MAIN}">'
            + "".join(f"<si><t>{s}</t></si>" for s in strings)
            + "</sst>",
        )
        zf.writestr(
            "xl/worksheets/sheet1.xml",
            f'<worksheet xmlns="{_STRICT_MAIN}"><sheetData>{body}</sheetData></worksheet>',
        )
    return str(path)


@pytest.mark.filterwarnings("ignore::UserWarning")  # openpyxl warns on its way to reading nothing
def test_openpyxl_reads_no_sheets_from_a_strict_workbook(tmp_path: Path) -> None:
    """The premise, asserted — otherwise the test below looks like belt-and-braces."""
    # Arrange
    path = _write_strict_xlsx(tmp_path, ["Barcode", GTIN13], [[0], [1]])

    # Act / Assert
    assert openpyxl.load_workbook(path, read_only=True).worksheets == []


def test_a_strict_workbook_reads_as_a_grid(tmp_path: Path) -> None:
    """The failure that actually bit: Strict OOXML *and* a title row above the table."""
    # Arrange: row 1 is a report title; the header is row 2.
    strings = ["Voorraadlijst Q3", "Artikelnr.", "Omschrijving NL", "Barcode", "A1", "Widget"]
    path = _write_strict_xlsx(tmp_path, strings, [[0], [1, 2, 3], [4, 5, None]])

    # Act
    sheet = read_process_list(_config(path))

    # Assert
    assert sheet.header == _HEADER
    assert sheet.rows == [["A1", "Widget", ""]]
    assert sheet.gtin_index == 2


def test_a_strict_workbook_yields_the_same_gtins_to_a_run(tmp_path: Path) -> None:
    # Arrange
    strings = ["Artikelnr.", "Omschrijving NL", "Barcode", "A1", "Widget", GTIN13]
    path = _write_strict_xlsx(tmp_path, strings, [[0, 1, 2], [3, 4, 5]])

    # Act / Assert
    assert load_process_list(_config(path)) == frozenset({GTIN14})
