"""Tests for ui/process_list_edit.py — installing, pruning and restoring the control file.

The operator's recurring job, and the one step where the shell writes to a file they authored.
The properties worth asserting are the ones that make that safe: other columns survive, the
previous version survives, the file they *uploaded* survives every save after it, and an empty
result is refused.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from lib.config import ProcessListConfig
from lib.errors import ProcessListError
from lib.process_list import load_process_list
from ui.process_list_edit import archive, archive_path, read_sheet, restore, save_sheet

GTIN_A = "8713195007359"
GTIN_B = "8713195007360"


def _write(tmp_path: Path, rows: list[list[object]], header: list[str] | None = None) -> Path:
    tmp_path.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(header or ["Artikelnr.", "Barcode", "Omschrijving"])
    for row in rows:
        sheet.append(row)
    path = tmp_path / "process-list.xlsx"
    workbook.save(path)
    return path


def _config(path: Path) -> ProcessListConfig:
    return ProcessListConfig(path=str(path), gtin_column="Barcode")


def test_reading_keeps_every_column(tmp_path: Path) -> None:
    """Only the GTIN column is configured; the rest are the operator's working notes."""
    path = _write(tmp_path, [["1079", GTIN_A, "Drain saver"]])

    sheet = read_sheet(_config(path))

    assert sheet.header == ["Artikelnr.", "Barcode", "Omschrijving"]
    assert sheet.rows == [["1079", GTIN_A, "Drain saver"]]
    assert sheet.gtin_index == 1


def test_an_integer_barcode_keeps_its_digits(tmp_path: Path) -> None:
    """openpyxl hands back a float for a numeric cell; ``8.7e+12`` is not a barcode."""
    path = _write(tmp_path, [["1079", int(GTIN_A), "Drain saver"]])

    sheet = read_sheet(_config(path))

    assert sheet.rows[0][1] == GTIN_A


def test_a_missing_gtin_column_says_so_the_way_the_cli_does(tmp_path: Path) -> None:
    path = _write(tmp_path, [["1079", GTIN_A]], header=["Artikelnr.", "EAN"])

    with pytest.raises(ProcessListError, match="Barcode"):
        read_sheet(_config(path))


def test_saving_keeps_the_other_columns_and_the_previous_version(tmp_path: Path) -> None:
    path = _write(tmp_path, [["1079", GTIN_A, "Drain saver"], ["1080", GTIN_B, "Airfryer basket"]])
    sheet = read_sheet(_config(path))

    backup = save_sheet(sheet.without({1}))

    assert backup.exists()  # there is no undo in a web form
    assert read_sheet(_config(backup)).rows[1][1] == GTIN_B  # the removed row is still recoverable
    kept = read_sheet(_config(path))
    assert kept.rows == [["1079", GTIN_A, "Drain saver"]]


def test_the_pruned_file_is_what_the_pipeline_then_reads(tmp_path: Path) -> None:
    """The point of the whole screen: the CLI must agree with what the operator just saved."""
    path = _write(tmp_path, [["1079", GTIN_A, "keep"], ["1080", GTIN_B, "drop"]])
    sheet = read_sheet(_config(path))

    save_sheet(sheet.without({1}))

    assert load_process_list(_config(path)) == frozenset({f"0{GTIN_A}"})


def test_saving_an_empty_list_is_refused(tmp_path: Path) -> None:
    """An empty control file yields an empty plan and a run that reports success publishing nothing.

    Refused here rather than at ``load_process_list``, because by then the operator's pruning is
    already lost and they have to redo it to find out.
    """
    path = _write(tmp_path, [["1079", GTIN_A, "Drain saver"]])
    sheet = read_sheet(_config(path))

    with pytest.raises(ProcessListError, match="report success"):
        save_sheet(sheet.without({0}))

    assert read_sheet(_config(path)).rows  # the file on disk is untouched


def test_without_does_not_mutate_the_original(tmp_path: Path) -> None:
    path = _write(tmp_path, [["1079", GTIN_A, "a"], ["1080", GTIN_B, "b"]])
    sheet = read_sheet(_config(path))

    pruned = sheet.without({0})

    assert len(sheet.rows) == 2
    assert len(pruned.rows) == 1


# --- Pruning in more than one pass --------------------------------------------
#
# The Data screen's grid keys each row by its position when the grid was built, and that key never
# changes. A sheet renumbers on every edit. Feeding fixed keys into a renumbered sheet is correct
# once and wrong from the second removal onward — and wrong in the way that matters least visibly:
# the grid shows one set of rows, the file receives another, and the save reports success. That is
# a live page and a permanent GS1 record for a product the operator did not choose.


def _rows(n: int) -> list[list[object]]:
    return [[f"art-{i}", f"871319500{i:04d}", f"name-{i}"] for i in range(n)]


def test_keeping_selects_by_original_position_and_holds_the_order(tmp_path: Path) -> None:
    sheet = read_sheet(_config(_write(tmp_path, _rows(5))))

    kept = sheet.keeping({0, 2, 4})

    assert [row[1] for row in kept.rows] == [sheet.rows[i][1] for i in (0, 2, 4)]
    assert len(sheet.rows) == 5, "the original is untouched"


def test_two_removals_leave_the_file_agreeing_with_the_grid(tmp_path: Path) -> None:
    """The regression. Remove one row, then another, exactly as the screen does it."""
    sheet = read_sheet(_config(_write(tmp_path, _rows(5))))
    grid = [{"_row": n, "gtin": row[1]} for n, row in enumerate(sheet.rows)]

    for selection in ({0}, {3}):  # two passes, keys taken from the original grid both times
        grid = [row for row in grid if row["_row"] not in selection]
        pruned = sheet.keeping({int(row["_row"]) for row in grid})

    assert [row[1] for row in pruned.rows] == [row["gtin"] for row in grid], (
        "the file would receive rows other than the ones left on screen"
    )


def test_the_incremental_form_is_the_one_that_drifts(tmp_path: Path) -> None:
    """Why ``keeping`` exists, asserted rather than described — ``without`` renumbers."""
    sheet = read_sheet(_config(_write(tmp_path, _rows(5))))
    grid = [{"_row": n, "gtin": row[1]} for n, row in enumerate(sheet.rows)]

    drifting = sheet
    for selection in ({0}, {3}):
        grid = [row for row in grid if row["_row"] not in selection]
        drifting = drifting.without(selection)

    assert [row[1] for row in drifting.rows] != [row["gtin"] for row in grid]


# --- The reader is the run's reader -------------------------------------------


def test_a_header_below_row_one_reads_on_screen_too(tmp_path: Path) -> None:
    """The bug this screen had: openpyxl, header fixed at row 1, against Strict-OOXML files.

    A real operator list has a report title above the table. It loaded in a run and failed here,
    which is the worst place for the two to disagree — the operator is looking at the screen.
    """
    # Arrange
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Voorraadlijst Q3"])
    worksheet.append([])
    worksheet.append(["Artikelnr.", "Barcode", "Omschrijving"])
    worksheet.append(["1079", GTIN_A, "Drain saver"])
    path = tmp_path / "process-list.xlsx"
    workbook.save(path)

    # Act
    sheet = read_sheet(_config(path))

    # Assert
    assert sheet.header == ["Artikelnr.", "Barcode", "Omschrijving"]
    assert sheet.rows == [["1079", GTIN_A, "Drain saver"]]
    assert sheet.gtin_index == 1


# --- The upload, and putting it back ------------------------------------------


def test_the_archive_sits_beside_the_control_file_and_is_not_the_backup(tmp_path: Path) -> None:
    """A collision would make Restore hand back the last pruned save, not the uploaded list."""
    # Arrange
    control = tmp_path / "process-list.xlsx"

    # Act
    kept = archive_path(control)

    # Assert
    assert kept == tmp_path / "process-list.source.xlsx"
    assert kept != control.with_suffix(".bak.xlsx")


def test_an_upload_is_archived_byte_for_byte_and_becomes_the_control_file(tmp_path: Path) -> None:
    # Arrange
    source = _write(tmp_path / "upload", [["1079", GTIN_A, "Drain saver"]])
    control = tmp_path / "input" / "process-list.xlsx"
    data = source.read_bytes()

    # Act
    kept = archive(_config(control), data)

    # Assert
    assert kept.read_bytes() == data
    assert control.read_bytes() == data
    assert load_process_list(_config(control)) == frozenset({f"0{GTIN_A}"})


def test_an_upload_that_will_not_read_is_refused_and_writes_nothing(tmp_path: Path) -> None:
    """Refused while the operator is still looking at the upload button, not on Preflight."""
    # Arrange
    control = _write(tmp_path, [["1079", GTIN_A, "Drain saver"]])
    before = control.read_bytes()

    # Act / Assert
    with pytest.raises(ProcessListError):
        archive(_config(control), b"this is not a workbook")

    assert control.read_bytes() == before, "the list they were working from is untouched"
    assert not archive_path(control).exists()


def test_an_upload_with_no_gtins_is_refused(tmp_path: Path) -> None:
    """An empty list plans nothing and reports success. Caught at the door."""
    # Arrange
    empty = _write(tmp_path / "upload", [["1079", None, "Drain saver"]])
    control = tmp_path / "input" / "process-list.xlsx"

    # Act / Assert
    with pytest.raises(ProcessListError, match="report success"):
        archive(_config(control), empty.read_bytes())

    assert not control.exists()


def test_restore_undoes_a_deselection(tmp_path: Path) -> None:
    """The whole reason deselection is safe: the worst case is one click."""
    # Arrange
    source = _write(tmp_path / "upload", [["1079", GTIN_A, "keep"], ["1080", GTIN_B, "drop"]])
    control = tmp_path / "input" / "process-list.xlsx"
    archive(_config(control), source.read_bytes())
    save_sheet(read_sheet(_config(control)).keeping({0}))
    assert load_process_list(_config(control)) == frozenset({f"0{GTIN_A}"})

    # Act
    backup = restore(_config(control))

    # Assert
    assert load_process_list(_config(control)) == frozenset({f"0{GTIN_A}", f"0{GTIN_B}"})
    assert read_sheet(_config(backup)).rows == [["1079", GTIN_A, "keep"]], (
        "the prune is recoverable"
    )


def test_the_archive_survives_saves_that_overwrite_the_backup(tmp_path: Path) -> None:
    """``.bak`` holds the previous save, so after two saves the uploaded list is only here."""
    # Arrange
    source = _write(
        tmp_path / "upload", [["1", GTIN_A, "a"], ["2", GTIN_B, "b"], ["3", "8713195007361", "c"]]
    )
    control = tmp_path / "input" / "process-list.xlsx"
    archive(_config(control), source.read_bytes())

    # Act: two prunes, keys taken from the original grid both times, as the screen does it.
    original = read_sheet(_config(control))
    save_sheet(original.keeping({0, 1}))
    save_sheet(original.keeping({0}))

    # Assert
    assert read_sheet(_config(control.with_suffix(".bak.xlsx"))).rows == [
        ["1", GTIN_A, "a"],
        ["2", GTIN_B, "b"],
    ], "the backup is one save old"
    restore(_config(control))
    assert len(read_sheet(_config(control)).rows) == 3, "the upload is still all three"


def test_restore_with_no_archive_raises(tmp_path: Path) -> None:
    """Doing nothing under a success message is the failure mode this project designs against."""
    # Arrange
    control = _write(tmp_path, [["1079", GTIN_A, "Drain saver"]])

    # Act / Assert
    with pytest.raises(ProcessListError, match="no uploaded scope list to restore"):
        restore(_config(control))


# --- What the saved file is like to open --------------------------------------


def test_the_saved_header_is_frozen_and_filterable(tmp_path: Path) -> None:
    """The file goes back to a spreadsheet; the operator's next act is to sort or filter it."""
    # Arrange
    path = _write(tmp_path, [["1079", GTIN_A, "Drain saver"], ["1080", GTIN_B, "Airfryer basket"]])

    # Act
    save_sheet(read_sheet(_config(path)).keeping({0, 1}))

    # Assert
    worksheet = openpyxl.load_workbook(path).active
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:C3"
