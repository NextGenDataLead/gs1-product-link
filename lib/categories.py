"""GPC brick → client category resolution (Phase 7.5).

The GDSN feed carries a product's GPC brick (``ProductRecord.gpc_brick_code``) but not the
client's marketing category. This module turns one into the other using the reviewed,
signed-off ``categories`` config (:class:`lib.config.CategoryConfig`): a per-GTIN override
wins, else the ``brick_category_map`` lookup, else nothing — an unmapped brick is *reported*
and leaves the category unset. The tool never guesses (a brick can span categories, so a
guess is a wrong page filing, not a near-miss).

Format-agnostic on purpose: it consumes :class:`~lib.records.ProductRecord`, not the GDSN
export, so it works for any export path. See ``docs/clients/noviplast-page-adapter.md`` §5.7.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass

import openpyxl

from lib.config import CategoryConfig
from lib.errors import ExportParseError
from lib.records import ProductRecord, SourceIssue, _coerce_cell

_log = logging.getLogger(__name__)

#: How many product names to show per brick when annotating a draft.
_DRAFT_SAMPLE_LIMIT: int = 3

#: How the ``category`` field is named where the operator finds it — the GPC brick, carried
#: on the ``BrickGPCCommercialData`` sheet as ``GpcCategoryCode``.
_CATEGORY_SOURCE: str = "GpcCategoryCode (GPC brick)"


@dataclass(frozen=True)
class CategoryResolution:
    """The outcome of resolving one product's category.

    Exactly one side is populated: a resolved ``term`` (``issue`` is ``None``), or ``term``
    ``None`` with an ``issue`` describing why it could not be resolved.
    """

    term: str | None
    issue: SourceIssue | None


def _normalise_overrides(overrides: dict[str, str]) -> dict[str, str]:
    """Key overrides by GTIN-14 so a 13-digit authored barcode matches a 14-digit product."""
    return {gtin.zfill(14): term for gtin, term in overrides.items()}


def _issue(product: ProductRecord, kind: str, detail: str) -> SourceIssue:
    """Build a ``category`` :class:`SourceIssue` for an unresolved product."""
    return SourceIssue(
        gtin=product.gtin,
        field="category",
        source=_CATEGORY_SOURCE,
        issue=kind,
        value=product.gpc_brick_code or "",
        detail=detail,
    )


def resolve_category(
    product: ProductRecord,
    *,
    brick_category_map: dict[str, str],
    overrides: dict[str, str],
    allowed_terms: frozenset[str],
) -> CategoryResolution:
    """Resolve a product's category term. Precedence: per-GTIN override > brick map > none.

    Args:
        product: The product to categorise.
        brick_category_map: GPC brick code → category term.
        overrides: GTIN → category term, winning over the brick map (matched on GTIN-14).
        allowed_terms: The closed set of permitted terms; a term outside it is treated as
            unresolved (defensive — the config validator already guarantees membership).

    Returns:
        A :class:`CategoryResolution`. An unmapped brick, an out-of-set term, or a product
        with no GPC brick yields ``term=None`` and a :class:`SourceIssue`. Never guesses.
    """
    override = _normalise_overrides(overrides).get(product.gtin14)
    if override is not None:
        if override in allowed_terms:
            return CategoryResolution(term=override, issue=None)
        return CategoryResolution(
            term=None,
            issue=_issue(
                product,
                "category_unmapped",
                f"per-GTIN override {override!r} is not an allowed category term",
            ),
        )

    brick = product.gpc_brick_code
    if brick is None:
        return CategoryResolution(
            term=None,
            issue=_issue(
                product, "category_brick_missing", "no GPC brick to derive a category from"
            ),
        )

    term = brick_category_map.get(brick)
    if term in allowed_terms:
        return CategoryResolution(term=term, issue=None)
    return CategoryResolution(
        term=None,
        issue=_issue(
            product,
            "category_unmapped",
            f"GPC brick {brick} maps to no category term — "
            "add it to brick_category_map or add a per-GTIN override",
        ),
    )


def distinct_bricks(products: list[ProductRecord]) -> dict[str, list[str]]:
    """Group the GTIN-14s in ``products`` by their GPC brick.

    Products without a brick are omitted (they have no brick to map); they surface instead
    as ``category_brick_missing`` findings from :func:`resolve_category`.
    """
    by_brick: dict[str, list[str]] = {}
    for product in products:
        if product.gpc_brick_code is not None:
            by_brick.setdefault(product.gpc_brick_code, []).append(product.gtin14)
    return by_brick


@dataclass(frozen=True)
class CoverageReport:
    """Which of the export's GPC bricks resolve to a term, and which do not (DoD #2).

    A brick counts as covered when it is in ``brick_category_map`` (``mapped``) or when every
    product carrying it has a per-GTIN override (``override_only``). ``unmapped`` maps each
    still-uncovered brick to the GTIN-14s under it that resolve to nothing.
    """

    total_bricks: int
    mapped: list[str]
    override_only: list[str]
    unmapped: dict[str, list[str]]

    @property
    def is_complete(self) -> bool:
        """Whether every brick in the export resolves to a term."""
        return not self.unmapped


# --- DIY datamodel parse + draft generation (DoD #1) -------------------------


#: How many rows to scan for the header before giving up on a worksheet. Real datamodels
#: carry a banner/numbering preamble above the header (the GS1 DIY workbook's is on row 4).
_HEADER_SCAN_LIMIT: int = 30


def _read_brick_sheet(
    worksheet: object, code_column: str, category_column: str
) -> dict[str, str] | None:
    """Read one worksheet into ``{brick_code: label}``, or ``None`` if it has no matching header.

    The header is the first row (within :data:`_HEADER_SCAN_LIMIT`) that carries both
    ``code_column`` and ``category_column``; rows below it are the data.
    """
    code_idx: int | None = None
    cat_idx = 0
    mapping: dict[str, str] = {}
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):  # type: ignore[attr-defined]
        if code_idx is None:
            if index >= _HEADER_SCAN_LIMIT:
                return None
            cells = [_coerce_cell(cell) for cell in row]
            if code_column in cells and category_column in cells:
                code_idx = cells.index(code_column)
                cat_idx = cells.index(category_column)
            continue
        code = _coerce_cell(row[code_idx]) if code_idx < len(row) else None
        label = _coerce_cell(row[cat_idx]) if cat_idx < len(row) else None
        if code:
            mapping[code] = label or ""
    return mapping if code_idx is not None else None


def load_diy_datamodel(
    path: str, *, code_column: str, category_column: str, sheet: str | None = None
) -> dict[str, str]:
    """Read the operator-supplied GS1 DIY sector datamodel into ``{brick_code: sector_label}``.

    The datamodel is supplied by the operator (like the export and control file), so its shape
    is not fixed here. The two column identities are parameters, the header may sit below a
    banner/preamble (found by scanning for a row that carries both columns), and the mapping may
    live on any worksheet: with ``sheet`` that one is read, otherwise every sheet is tried and
    the first with a matching header wins.

    Args:
        path: Path to the datamodel workbook.
        code_column: Header of the column holding the GPC brick code.
        category_column: Header of the column holding the DIY sector label.
        sheet: Worksheet to read; ``None`` scans all sheets for a matching header.

    Returns:
        Mapping of brick code to its DIY sector label (label ``""`` when the cell is blank).

    Raises:
        ExportParseError: If the file cannot be read, or no worksheet carries both columns.
    """
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except (FileNotFoundError, OSError) as exc:
        raise ExportParseError(f"cannot read DIY datamodel at {path}: {exc}") from exc
    try:
        names = [sheet] if sheet is not None else list(workbook.sheetnames)
        if sheet is not None and sheet not in workbook.sheetnames:
            raise ExportParseError(f"DIY datamodel at {path}: no worksheet named {sheet!r}")
        for name in names:
            mapping = _read_brick_sheet(workbook[name], code_column, category_column)
            if mapping is not None:
                return mapping
    finally:
        workbook.close()
    raise ExportParseError(
        f"DIY datamodel at {path}: no worksheet has a header row with both "
        f"{code_column!r} and {category_column!r}"
    )


@dataclass(frozen=True)
class BrickMapDraft:
    """A human-review skeleton for ``brick_category_map`` (DoD #1).

    Every brick in the export is present with an UNSET (``""``) term for a person to fill; the
    datamodel supplies the DIY *sector* label, not the client's site term, so the mapping from
    label to term — and the client sign-off — remain a human step. ``annotations`` gives each
    brick its DIY label, product count, and sample names; ``unannotated`` lists bricks the
    datamodel did not cover.
    """

    entries: dict[str, str]
    annotations: dict[str, str]
    unannotated: list[str]


def _names_by_gtin14(products: list[ProductRecord]) -> dict[str, str]:
    """First available product name per GTIN-14, for annotating a draft."""
    names: dict[str, str] = {}
    for product in products:
        values = product.product_name.values
        if values:
            names[product.gtin14] = next(iter(values.values()))
    return names


def draft_brick_map(
    bricks: dict[str, list[str]],
    products: list[ProductRecord],
    datamodel: dict[str, str] | None,
) -> BrickMapDraft:
    """Build a review skeleton covering every brick in ``bricks`` (from :func:`distinct_bricks`).

    Args:
        bricks: Brick code → GTIN-14s carrying it.
        products: The products, used to annotate each brick with sample names.
        datamodel: Optional brick → DIY sector label from :func:`load_diy_datamodel`.

    Returns:
        A :class:`BrickMapDraft` with every brick UNSET, annotated, and the datamodel's
        coverage gap recorded.
    """
    names = _names_by_gtin14(products)
    entries: dict[str, str] = {}
    annotations: dict[str, str] = {}
    unannotated: list[str] = []
    for brick in sorted(bricks):
        gtins = bricks[brick]
        entries[brick] = ""
        label = datamodel.get(brick) if datamodel else None
        if not label:
            unannotated.append(brick)
        samples: list[str] = []
        for gtin in gtins:
            name = names.get(gtin)
            if name and name not in samples:
                samples.append(name)
            if len(samples) >= _DRAFT_SAMPLE_LIMIT:
                break
        parts = [label] if label else []
        parts.append(f"{len(gtins)} product(s)")
        if samples:
            parts.append("e.g. " + ", ".join(samples))
        annotations[brick] = " | ".join(parts)
    return BrickMapDraft(entries=entries, annotations=annotations, unannotated=unannotated)


def coverage_report(products: list[ProductRecord], categories: CategoryConfig) -> CoverageReport:
    """Verify every GPC brick in ``products`` resolves to a term under ``categories``.

    Iterates products (not just bricks) so a brick covered entirely by per-GTIN overrides
    counts as covered. This is the machine-checkable form of "every brick maps to a category".
    """
    allowed = frozenset(categories.terms)
    by_brick: dict[str, list[ProductRecord]] = {}
    for product in products:
        if product.gpc_brick_code is not None:
            by_brick.setdefault(product.gpc_brick_code, []).append(product)

    mapped: list[str] = []
    override_only: list[str] = []
    unmapped: dict[str, list[str]] = {}
    for brick, brick_products in by_brick.items():
        if brick in categories.brick_category_map:
            mapped.append(brick)
            continue
        unresolved = [
            product.gtin14
            for product in brick_products
            if resolve_category(
                product,
                brick_category_map=categories.brick_category_map,
                overrides=categories.overrides,
                allowed_terms=allowed,
            ).term
            is None
        ]
        if unresolved:
            unmapped[brick] = unresolved
        else:
            override_only.append(brick)

    return CoverageReport(
        total_bricks=len(by_brick),
        mapped=sorted(mapped),
        override_only=sorted(override_only),
        unmapped=unmapped,
    )
