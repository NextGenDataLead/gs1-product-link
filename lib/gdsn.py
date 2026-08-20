"""Reader for GS1 Data Source / GDSN datapool Excel exports.

This is a spec extension (see ``docs/IMPLEMENTATION_SPEC.md`` §3 notes). Where
``lib.records.parse_excel_row`` handles a flat single-sheet export, this module
handles the rich multi-worksheet GDSN datapool export the pilot client (Democlient)
actually produces.

Structure of a GDSN export:

* One worksheet per GDSN module (``TradeItemDescription``, ``MarketingInformation``,
  ``TradeItemMeasurements``, ``ReferencedFileDetailInformation``, ...).
* Seven header rows per sheet (data starts on the eighth). Each column's identity is
  a nested attribute *path* (e.g. ``TradeItemDescriptionInformation > DescriptionShort[0]
  > Value``) plus a human *label* carrying the stable GDSN attribute number, e.g.
  ``"Short product name (3297)"``.
* Every sheet is keyed on ``Gtin`` + ``TargetMarketCountryCode`` +
  ``TradeItemUnitDescriptorCode``; the same GTIN recurs once per target market.
* Localised text is stored as adjacent ``LanguageCode`` / ``Value`` column pairs within
  a repeated group; measurements as ``MeasurementUnitCode`` / ``Value`` pairs.

The client declares, in ``clients.yml``, which GDSN attribute feeds each
:class:`~lib.records.ProductRecord` field (a :class:`GdsnSource`) and the order to
consult markets (``market_priority``). :func:`build_records` joins across sheets by GTIN,
takes the first non-blank value per field/language walking that order, and reports the
gaps and cross-market disagreements it finds along the way.
"""

from __future__ import annotations

import logging
import re
from collections.abc import Callable
from dataclasses import dataclass
from difflib import SequenceMatcher
from typing import Final, NamedTuple

import openpyxl
from pydantic import BaseModel, ConfigDict, model_validator

from lib.errors import ExportParseError
from lib.records import ProductRecord, SourceIssue, _coerce_cell, build_product_record

_log = logging.getLogger(__name__)

# --- Constants ---------------------------------------------------------------

#: The four key columns present at the start of every GDSN module sheet.
_GTIN_SEGMENT: Final = "Gtin"
_MARKET_SEGMENT: Final = "TargetMarketCountryCode"
_UNIT_SEGMENT: Final = "TradeItemUnitDescriptorCode"

#: Only consumer base units carry the product-page content we publish.
CONSUMER_UNIT: Final = "BASE_UNIT_OR_EACH"

#: Header leaf names with special pairing/interpretation semantics.
_LEAF_VALUE: Final = "Value"
_LEAF_LANGUAGE: Final = "LanguageCode"
_LEAF_UNIT: Final = "MeasurementUnitCode"
_LEAF_URI: Final = "UniformResourceIdentifier"
_LEAF_IS_PRIMARY: Final = "IsPrimaryFile"
_FILE_GROUP_PREFIX: Final = "ReferencedFileHeader"
_FILE_TYPE_SEGMENT: Final = "ReferencedFileTypeCode"
_PRODUCT_IMAGE_TYPE: Final = "PRODUCT_IMAGE"

#: ProductRecord fields that must resolve for a record to be publishable (E5/E17).
REQUIRED_FIELDS: Final[frozenset[str]] = frozenset({"brand", "product_name"})

#: Truthy spellings of the ``IsPrimaryFile`` flag.
_TRUE_VALUES: Final[frozenset[str]] = frozenset({"true", "1", "yes"})

#: How far to scan for the first data row before giving up on a sheet.
_MAX_HEADER_SCAN: Final = 40

_ATTR_ID_RE: Final = re.compile(r"\((\d+)\)\s*$")
_INDEX_RE: Final = re.compile(r"\[\d+\]$")

#: How a language-agnostic ``multivalue`` attribute's repeated slots are joined. Not a newline:
#: the value renders verbatim inside a page's Technische details, where HTML would collapse one
#: to a space and turn ``kunststof metaal`` into what reads as a single material.
SCALAR_SEPARATOR: Final = ", "


def _strip_index(segment: str) -> str:
    """Drop a trailing repeated-group index, e.g. ``"DescriptionShort[0]"`` → base."""
    return _INDEX_RE.sub("", segment)


# --- Mapping source (declared per client in clients.yml) ---------------------


class GdsnSource(BaseModel):
    """Where a :class:`~lib.records.ProductRecord` field is sourced in a GDSN export.

    Attributes:
        sheet: The worksheet (GDSN module) holding the attribute.
        attribute: The GDSN attribute number (e.g. ``"3297"``) or a path segment
            name (e.g. ``"GpcCategoryCode"``) identifying the column.
        localised: Whether the attribute is a per-language ``LanguageCode``/``Value``
            group (resolved once per configured language).
        multivalue: For an attribute that repeats across ``[n]`` slots (e.g. every
            ``TradeItemFeatureBenefit[n]``, or ``Material[0..2]``), read every slot instead of
            only the first. Blank slots are skipped, so a hole in the sequence does not become
            an empty item. **The separator differs by kind, and deliberately:** a ``localised``
            attribute joins with a newline, because the generator splits attr 1067 back into
            ranked USP candidates on exactly that character; a language-agnostic one joins with
            :data:`SCALAR_SEPARATOR`, because its value renders verbatim in a page's Technische
            details, where a newline would collapse to a space and read as one fused word.
        with_unit: Whether to append the paired ``MeasurementUnitCode`` to the value.
        primary_file: Whether to resolve the primary referenced-file URI instead of a
            plain attribute (used for ``image_url``).
        strip_prefix: A literal prefix to remove from the resolved value when present
            (e.g. ``"Noviplast "``, where the feed repeats the brand in the product name
            but the page renders brand separately). Matched **exactly**: a value that only
            resembles the prefix is left untouched and reported, never corrected — see
            :func:`_strip_prefix`.
        max_length: The longest this value is expected to be, in characters. A longer value
            is **reported and kept**, never truncated — see :func:`_check_length`. ``0``
            (the default) means no expectation.
        report_issues: Whether this field's cross-market source-quality findings —
            ``value_blank`` (absent from every market) and ``value_inconsistent_across_markets``
            (markets disagree) — reach ``source_issues.json``. Default ``True`` — a
            published field's gaps and conflicts are the operator's work queue. Set
            ``False`` for a field the tool parses but does not publish directly, e.g. a
            generator *input*: its findings are the generator's future work, not today's
            source-fix queue, and surfacing them now only asks the operator about a field
            they cannot see on the page. Does not affect ``value_too_long`` /
            ``brand_prefix_mismatch``, which are gated by ``max_length`` / ``strip_prefix``.
        required: Whether a product missing this value may never publish (E23). A localised
            field must be non-blank in **every** configured language: the hold is per *product*,
            not per language, so a SKU is never half-published — see
            :func:`lib.mandatory.missing_mandatory`.
        required_group: Name of an either-or group. A product satisfies the group when **at
            least one** member carries a value, so alternatives can be declared without making
            each one individually mandatory. Noviplast's marketing copy is the case this exists
            for: the generator can write from ``1083`` *or* ``1067``, and needs only one of them.
            Mutually exclusive with ``required`` — a field is individually mandatory or part of
            a group, never both.
        in_matrix: Whether this field gets a column in the data-quality report's coverage matrix.
            Default ``True``. Set ``False`` for a value the pipeline carries but nothing consumes:
            a column that is present for every product and feeds nothing is noise in a table whose
            job is to show where the gaps are. It does **not** stop the field being parsed — the
            value is still in ``extras`` for whatever future check wants it.
        translate: Whether the generator may fill this value in a language the feed lacks, by
            translating the same value from a language the feed has. Default ``False`` —
            **opt in per field**, because every filled value costs producer tokens and adds
            LLM-written text to a page. Deriving a French value from the Dutch one the feed
            already carries is translation, not invention; a field blank in **every** language is
            never filled, and stays a source finding (E23). Each filled value is reported for
            the operator to put back into MyGS1 — see :func:`lib.generator.translation_gaps`.
    """

    model_config = ConfigDict(frozen=True)

    sheet: str
    attribute: str = ""
    localised: bool = False
    multivalue: bool = False
    with_unit: bool = False
    primary_file: bool = False
    strip_prefix: str = ""
    max_length: int = 0
    report_issues: bool = True
    required: bool = False
    required_group: str = ""
    in_matrix: bool = True
    translate: bool = False

    @model_validator(mode="after")
    def _required_xor_group(self) -> GdsnSource:
        """``required`` and ``required_group`` answer different questions; naming both is a bug."""
        if self.required and self.required_group:
            raise ValueError(
                "a gdsn_map field is either individually required or part of a required_group, "
                "not both"
            )
        return self


# --- Column / sheet models ---------------------------------------------------


def is_mandatory(source: GdsnSource) -> bool:
    """Whether a product missing this value may never publish (E23).

    One predicate for the two ways a source can be mandatory, because "individually required" and
    "the only member of its either-or group carrying a value" are the same fact to every reader:
    :func:`lib.mandatory.missing_mandatory`, which holds the SKU, and the coverage matrix, which
    tells the operator that it will. Those two must never disagree about which columns matter.
    """
    return bool(source.required or source.required_group)


@dataclass(frozen=True)
class GdsnColumn:
    """One resolved column in a GDSN sheet."""

    index: int
    path: tuple[str, ...]
    label: str | None
    attr_id: str | None

    @property
    def leaf_name(self) -> str:
        """The index-stripped final path segment (e.g. ``"Value"``)."""
        return _strip_index(self.path[-1]) if self.path else ""

    @property
    def group_path(self) -> tuple[str, ...]:
        """The path with its leaf removed, used to pair sibling columns."""
        return self.path[:-1]

    def matches_attribute(self, attribute: str) -> bool:
        """Whether this column belongs to ``attribute`` (number or segment name)."""
        if attribute.isdigit():
            return self.attr_id == attribute
        return any(_strip_index(seg) == attribute for seg in self.path)


class GdsnSheet:
    """A parsed GDSN worksheet: its columns plus a (gtin, market) → row index."""

    def __init__(
        self,
        name: str,
        columns: list[GdsnColumn],
        rows_by_key: dict[tuple[str, str], tuple[object, ...]],
    ) -> None:
        self.name = name
        self.columns = columns
        self.rows_by_key = rows_by_key

    def _cell(self, row: tuple[object, ...], index: int) -> str | None:
        return _coerce_cell(row[index]) if index < len(row) else None

    def _sibling(self, group_path: tuple[str, ...], leaf: str) -> GdsnColumn | None:
        return next(
            (c for c in self.columns if c.group_path == group_path and c.leaf_name == leaf),
            None,
        )

    def has_attribute(self, attribute: str) -> bool:
        """Whether any column resolves the given attribute."""
        return any(c.matches_attribute(attribute) for c in self.columns)

    def pick_localised(self, gtin: str, market: str, attribute: str, lang: str) -> str | None:
        """Return the ``Value`` whose paired ``LanguageCode`` matches ``lang``."""
        row = self.rows_by_key.get((gtin, market))
        if row is None:
            return None
        for value_col in self.columns:
            if value_col.leaf_name != _LEAF_VALUE or not value_col.matches_attribute(attribute):
                continue
            lang_col = self._sibling(value_col.group_path, _LEAF_LANGUAGE)
            if lang_col is None:
                continue
            cell_lang = self._cell(row, lang_col.index)
            if cell_lang and cell_lang.strip().lower() == lang.lower():
                value = self._cell(row, value_col.index)
                if value:
                    return value
        return None

    def pick_localised_all(self, gtin: str, market: str, attribute: str, lang: str) -> list[str]:
        """Return every ``Value`` matching ``lang`` for ``attribute``, in column (slot) order.

        Unlike :meth:`pick_localised`, which returns the first match, this collects every repeated
        group — e.g. all ``TradeItemFeatureBenefit[n]`` slots — so a multi-slot attribute is
        captured whole rather than truncated to its first slot.
        """
        row = self.rows_by_key.get((gtin, market))
        if row is None:
            return []
        values: list[str] = []
        for value_col in self.columns:
            if value_col.leaf_name != _LEAF_VALUE or not value_col.matches_attribute(attribute):
                continue
            lang_col = self._sibling(value_col.group_path, _LEAF_LANGUAGE)
            if lang_col is None:
                continue
            cell_lang = self._cell(row, lang_col.index)
            if cell_lang and cell_lang.strip().lower() == lang.lower():
                value = self._cell(row, value_col.index)
                if value:
                    values.append(value)
        return values

    def _scalar_value_columns(self, attribute: str) -> list[GdsnColumn]:
        """The columns holding ``attribute``'s own value, in sheet (slot) order.

        Two regimes, because a GDSN attribute is spelled two ways: a ``LanguageCode``/``Value``
        group contributes its ``Value`` leaves, and a bare column (``GpcCategoryCode``) is its own
        value. The second is only consulted when the first finds nothing, so a group's unit and
        language siblings can never be mistaken for the value.

        That preference is **unfalsifiable on this feed** — mutating it away changes nothing, in
        the suite or across all 127 real products, because no group here carries a ``Value`` leaf
        *and* another non-unit, non-language leaf under the same attribute. It is kept because it
        is what the single-value picker this was extracted from already did, and because the
        fallback exists precisely for shapes this export does not have.
        """
        candidates = [c for c in self.columns if c.matches_attribute(attribute)]
        values = [c for c in candidates if c.leaf_name == _LEAF_VALUE]
        if values:
            return values
        return [c for c in candidates if c.leaf_name not in (_LEAF_LANGUAGE, _LEAF_UNIT)]

    def _scalar_cell(
        self, row: tuple[object, ...], value_col: GdsnColumn, with_unit: bool
    ) -> str | None:
        """One slot's value, with its own group's unit code appended when asked for.

        The unit is read from ``value_col``'s own ``group_path``: a repeated attribute carries one
        ``MeasurementUnitCode`` per slot, and borrowing the first slot's would report the second
        slot's number in the wrong unit.
        """
        value = self._cell(row, value_col.index)
        if value is None:
            return None
        if with_unit:
            unit_col = self._sibling(value_col.group_path, _LEAF_UNIT)
            unit = self._cell(row, unit_col.index) if unit_col else None
            if unit:
                return f"{value} {unit}"
        return value

    def pick_scalar(
        self, gtin: str, market: str, attribute: str, with_unit: bool = False
    ) -> str | None:
        """Return a language-agnostic attribute value, optionally with its unit.

        The **first** slot only, blank included: a repeated attribute whose slot 0 is empty reads
        as empty here, and does not quietly fall through to slot 1. Only a source that opts into
        ``multivalue`` goes on to :meth:`pick_scalar_all`.
        """
        row = self.rows_by_key.get((gtin, market))
        if row is None:
            return None
        value_col = next(iter(self._scalar_value_columns(attribute)), None)
        if value_col is None:
            return None
        return self._scalar_cell(row, value_col, with_unit)

    def pick_scalar_all(
        self, gtin: str, market: str, attribute: str, with_unit: bool = False
    ) -> list[str]:
        """Return every slot's value for a language-agnostic ``attribute``, in slot order.

        The language-agnostic twin of :meth:`pick_localised_all`, down to its truthiness test:
        :func:`lib.records._coerce_cell` already turns ``None``, ``""`` and a whitespace-only cell
        all into ``None``, so no falsy-but-present value can reach here and ``if value`` cannot be
        told from ``if value is not None``. Belt and braces against a change to that coercion, not
        live behaviour — a mutation of it survives, and is meant to.

        Blank slots are skipped rather than yielded as empty strings, so a hole in the middle of
        ``Material[0..2]`` joins to
        ``"kunststof, glas"`` and not to ``"kunststof, , glas"``.
        """
        row = self.rows_by_key.get((gtin, market))
        if row is None:
            return []
        values: list[str] = []
        for value_col in self._scalar_value_columns(attribute):
            value = self._scalar_cell(row, value_col, with_unit)
            if value:
                values.append(value)
        return values

    def pick_primary_file(self, gtin: str, market: str) -> str | None:
        """Return the URI of the primary product image, with graceful fallbacks."""
        row = self.rows_by_key.get((gtin, market))
        if row is None:
            return None
        groups: dict[str, list[GdsnColumn]] = {}
        for col in self.columns:
            if col.path and col.path[0].startswith(_FILE_GROUP_PREFIX):
                groups.setdefault(col.path[0], []).append(col)

        primary_uris: list[str] = []
        image_uris: list[str] = []
        all_uris: list[str] = []
        for cols in groups.values():
            uri = self._leaf_cell(row, cols, _LEAF_URI)
            if not uri:
                continue
            all_uris.append(uri)
            primary_flag = (self._leaf_cell(row, cols, _LEAF_IS_PRIMARY) or "").lower()
            is_primary = primary_flag in _TRUE_VALUES
            file_type = self._type_cell(row, cols)
            if is_primary:
                primary_uris.append(uri)
            if file_type == _PRODUCT_IMAGE_TYPE:
                image_uris.append(uri)
        for bucket in (primary_uris, image_uris, all_uris):
            if bucket:
                return bucket[0]
        return None

    def _leaf_cell(self, row: tuple[object, ...], cols: list[GdsnColumn], leaf: str) -> str | None:
        col = next((c for c in cols if c.leaf_name == leaf), None)
        return self._cell(row, col.index) if col else None

    def _type_cell(self, row: tuple[object, ...], cols: list[GdsnColumn]) -> str | None:
        col = next(
            (
                c
                for c in cols
                if c.leaf_name == _LEAF_VALUE and any(_FILE_TYPE_SEGMENT in seg for seg in c.path)
            ),
            None,
        )
        return self._cell(row, col.index) if col else None


# --- Workbook parsing --------------------------------------------------------


def _data_start_row(rows: list[tuple[object, ...]]) -> int | None:
    """Index of the first row whose first cell is an all-digit GTIN."""
    for i, row in enumerate(rows[:_MAX_HEADER_SCAN]):
        first = row[0] if row else None
        if first is not None and str(first).strip().isdigit():
            return i
    return None


def _parse_columns(rows: list[tuple[object, ...]], data_start: int) -> list[GdsnColumn]:
    """Reconstruct each column's attribute path and label from the header rows."""
    header_rows = rows[:data_start]
    ncols = max((len(r) for r in header_rows), default=0)
    columns: list[GdsnColumn] = []
    for idx in range(ncols):
        segments = [
            str(hr[idx]).strip()
            for hr in header_rows
            if idx < len(hr) and hr[idx] not in (None, "")
        ]
        if not segments:
            columns.append(GdsnColumn(index=idx, path=(), label=None, attr_id=None))
            continue
        path: tuple[str, ...]
        if len(segments) == 1:
            path, label = (segments[0],), segments[0]
        else:
            path, label = tuple(segments[:-1]), segments[-1]
        match = _ATTR_ID_RE.search(label)
        columns.append(
            GdsnColumn(index=idx, path=path, label=label, attr_id=match.group(1) if match else None)
        )
    return columns


def _key_index(columns: list[GdsnColumn], segment: str, fallback: int) -> int:
    """Locate a key column by its first path segment, else use the positional fallback."""
    for col in columns:
        if col.path and col.path[0] == segment:
            return col.index
    return fallback


def read_workbook(path: str) -> dict[str, GdsnSheet]:
    """Parse a GDSN datapool workbook into per-sheet models.

    Args:
        path: Filesystem path to the ``.xlsx`` export.

    Returns:
        Mapping of sheet name to :class:`GdsnSheet`. Reference/metadata sheets with
        no digit-keyed data rows are skipped.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: dict[str, GdsnSheet] = {}
        for name in workbook.sheetnames:
            rows = list(workbook[name].iter_rows(values_only=True))
            data_start = _data_start_row(rows)
            if data_start is None:
                continue
            columns = _parse_columns(rows, data_start)
            gtin_idx = _key_index(columns, _GTIN_SEGMENT, 0)
            market_idx = _key_index(columns, _MARKET_SEGMENT, 1)
            unit_idx = _key_index(columns, _UNIT_SEGMENT, 3)
            rows_by_key: dict[tuple[str, str], tuple[object, ...]] = {}
            for row in rows[data_start:]:
                gtin = _coerce_cell(row[gtin_idx]) if gtin_idx < len(row) else None
                if gtin is None or not gtin.isdigit():
                    continue  # E4: empty / key-less row skipped silently
                unit = _coerce_cell(row[unit_idx]) if unit_idx < len(row) else None
                if unit is not None and unit != CONSUMER_UNIT:
                    continue
                market = _coerce_cell(row[market_idx]) if market_idx < len(row) else None
                rows_by_key[(gtin, market or "")] = row
            sheets[name] = GdsnSheet(name=name, columns=columns, rows_by_key=rows_by_key)
        return sheets
    finally:
        workbook.close()


# --- Record building ---------------------------------------------------------


class BuildResult(NamedTuple):
    """Outcome of :func:`build_records`: records plus non-fatal/fatal messages.

    ``issues`` is the structured subset of ``warnings`` that names a defect in the *source
    datapool* rather than in the tool's config or this run — the things a person has to go
    and fix in MyGS1. They are reported both ways on purpose: as a warning so the run's
    summary counts them, and as a :class:`~lib.records.SourceIssue` so
    ``scripts/parse_export.py`` can write them to a file that outlives the terminal.
    """

    records: list[ProductRecord]
    warnings: list[str]
    errors: list[str]
    issues: list[SourceIssue] = []  # noqa: RUF012 — NamedTuple default, never mutated in place


def _validate_sources(
    workbook: dict[str, GdsnSheet],
    gdsn_map: dict[str, GdsnSource],
) -> tuple[list[str], list[str]]:
    """Check each mapped source resolves; required fields error, optional fields warn (E17)."""
    warnings: list[str] = []
    errors: list[str] = []
    for field, src in gdsn_map.items():
        if field == "gtin":
            continue
        sheet = workbook.get(src.sheet)
        missing = sheet is None or (
            not src.primary_file and bool(src.attribute) and not sheet.has_attribute(src.attribute)
        )
        if not missing:
            continue
        detail = (
            f"field {field!r}: source sheet {src.sheet!r}/attribute {src.attribute!r} not found"
        )
        (errors if field in REQUIRED_FIELDS else warnings).append(detail)
    return warnings, errors


def build_records(  # noqa: PLR0913 — each argument is a distinct input; bundling hides them
    workbook: dict[str, GdsnSheet],
    gdsn_map: dict[str, GdsnSource],
    market_priority: list[str],
    languages: list[str],
    default_language: str,
    gdsn_extras: dict[str, GdsnSource] | None = None,
) -> BuildResult:
    """Join GDSN sheets by GTIN into canonical :class:`~lib.records.ProductRecord`s.

    Args:
        workbook: Parsed sheets from :func:`read_workbook`.
        gdsn_map: ProductRecord field → :class:`GdsnSource`.
        market_priority: Market codes in the order to consult them (e.g.
            ``["528", "056", "276", "442"]``). For each field and language the first
            market that supplies a non-blank value wins; the same list picks scalars.
            Replaces the old ``{market: language}`` map, which baked in a 1:1
            market↔language constraint the real export contradicts — every market row
            carries every language, so which market *has* a given value varies by product.
        languages: The languages to resolve per product (the site's ``wordpress.languages``).
            No longer derivable from the market map, so passed explicitly.
        default_language: The language whose ``product_name`` is required (E5); must be in
            ``languages``.
        gdsn_extras: Optional named pass-through attributes carried into ``extras``.

    Returns:
        A :class:`BuildResult`. ``records`` holds successfully built products; ``errors``
        holds per-GTIN or config-level failures (caller exits non-zero and writes nothing).
    """
    gdsn_extras = gdsn_extras or {}
    warnings, errors = _validate_sources(workbook, gdsn_map)
    if errors:
        return BuildResult(records=[], warnings=warnings, errors=errors)

    if not market_priority:
        return BuildResult(records=[], warnings=warnings, errors=["market_priority is empty"])
    if default_language not in languages:
        return BuildResult(
            records=[],
            warnings=warnings,
            errors=[f"default_language {default_language!r} not in languages {languages}"],
        )
    ctx = _BuildContext(
        workbook=workbook,
        market_priority=market_priority,
        languages=languages,
        default_language=default_language,
    )

    gtins = sorted({gtin for sheet in workbook.values() for (gtin, _market) in sheet.rows_by_key})
    records: list[ProductRecord] = []
    issues: list[SourceIssue] = []
    for gtin in gtins:
        acc = _Accumulator(
            scalars={"gtin": gtin},
            localised={},
            extras={},
            extras_localised={},
            warnings=[],
            issues=[],
        )
        for field, src in gdsn_map.items():
            if field != "gtin":
                _resolve_field(ctx, field, src, gtin, acc)
        _resolve_extras(ctx, gdsn_extras, gtin, acc)
        warnings.extend(acc.warnings)
        issues.extend(acc.issues)

        product_name = acc.localised.get("product_name")
        if not product_name or default_language not in product_name:
            errors.append(f"GTIN {gtin}: missing product_name.{default_language}")  # E5
            continue
        if "product_name" in gdsn_map:
            issues.extend(
                _check_field_language(
                    product_name, "product_name", source_label(gdsn_map["product_name"]), gtin
                )
            )
        try:
            records.append(
                build_product_record(
                    gtin=gtin,
                    scalars=acc.scalars,
                    localised=acc.localised,
                    extras=acc.extras,
                    extras_localised=acc.extras_localised,
                )
            )
        except ExportParseError as exc:
            errors.append(str(exc))
    return BuildResult(records=records, warnings=warnings, errors=errors, issues=issues)


@dataclass(frozen=True)
class _BuildContext:
    """Shared inputs threaded through per-field resolution."""

    workbook: dict[str, GdsnSheet]
    #: Markets in the order to consult them; the first with a value wins (per GTIN, per
    #: field, per language). Replaces the old static ``lang_to_market``/``primary_market``
    #: pair — with every market carrying every language, the market that actually holds a
    #: value varies by product, so resolution walks the ranking rather than reading a map.
    market_priority: list[str]
    languages: list[str]
    default_language: str


@dataclass
class _Accumulator:
    """Per-GTIN field values collected before constructing the record."""

    scalars: dict[str, str]
    localised: dict[str, dict[str, str]]
    extras: dict[str, str]
    #: Pass-through extras whose source attribute is per-language, keyed name → language.
    extras_localised: dict[str, dict[str, str]]
    #: Non-fatal notes raised while resolving this GTIN's fields; merged into
    #: :attr:`BuildResult.warnings` so ``parse_export``'s summary counts them.
    warnings: list[str]
    #: The structured form of those notes that name a source-datapool defect; merged into
    #: :attr:`BuildResult.issues` and written to ``data/source_issues.json``.
    issues: list[SourceIssue]


#: How close a value's opening must be to ``strip_prefix`` before it is reported as a likely
#: misspelling. Tuned against the pilot export: the real typos ("Noviplat", "Nociplast",
#: "Novilplast" for "Noviplast ") score ~0.94, while genuinely unprefixed names ("Super Glove",
#: "Plasma Lighter", "Garden Clipper") score below 0.4 — a wide gap, so the threshold is not
#: delicate.
_PREFIX_TYPO_RATIO: Final = 0.8


class _Where(NamedTuple):
    """Where a reported value came from, in both vocabularies (see :class:`SourceIssue`)."""

    field: str  #: ours, e.g. product_name.nl
    source: str  #: the source system's, e.g. TradeItemDescription attr 3318
    gtin: str


def _strip_prefix(value: str, prefix: str, where: _Where, acc: _Accumulator) -> str:
    """Remove ``prefix`` from ``value`` when it matches exactly; report near-misses (§4.1).

    Args:
        value: The resolved field value.
        prefix: The literal prefix to remove.
        where: Field identity for the report, in both vocabularies.
        acc: Collector for this GTIN's notes; a near-miss lands in both its ``warnings``
            (so the run summary counts it) and its ``issues`` (so it reaches the file).

    The prefix is matched literally and never repaired. A value whose opening merely
    *resembles* the prefix is a defect in the source datapool — a misspelled or unspaced
    brand — and correcting it here would hide the defect while the wrong text stays
    authoritative upstream. So it is reported and passed through unchanged, and the operator
    fixes it at source (the same principle as the generated-content report: surface the gap,
    do not paper over it).

    The note goes to ``warnings`` as well as the log, so ``parse_export``'s summary counts
    it — a warning the summary reports as "0 warnings" is one nobody acts on.

    Genuinely unprefixed values are silent: not every product name repeats the brand.
    """
    if value.startswith(prefix):
        return value[len(prefix) :].lstrip()
    opening = value[: len(prefix)]
    if SequenceMatcher(None, opening.casefold(), prefix.casefold()).ratio() >= _PREFIX_TYPO_RATIO:
        detail = (
            f"starts with {opening!r}, which resembles but does not match the configured "
            f"strip_prefix {prefix!r} — likely a misspelling in the source data; left "
            f"unchanged, fix it at the source"
        )
        acc.warnings.append(f"{where.field} for {where.gtin} {detail}")
        acc.issues.append(
            SourceIssue(
                gtin=where.gtin,
                field=where.field,
                source=where.source,
                issue="brand_prefix_mismatch",
                value=value,
                detail=detail,
            )
        )
        _log.warning("%s for %s %s", where.field, where.gtin, detail)
    return value


def _check_length(value: str, limit: int, where: _Where, acc: _Accumulator) -> str:
    """Report a value longer than ``limit``; return it **unchanged** (§4.2).

    Kept rather than truncated, for the same reason a near-miss prefix is not repaired: the
    datapool is authoritative, and a value silently cut here stays too long in MyGS1 and
    returns on the next export — while the page shows a sentence severed mid-word.

    The case this exists for: Democlient's tagline is mapped to GS1 attr 1083
    *TradeItemMarketingMessage*, which is free-text marketing copy **by definition** and is
    frequently a paragraph (fr median 150 chars, max 1433) where the page's tagline slot
    wants one line (~31 on live pages). That is a mapping mismatch, not a typo, so the fix
    is upstream — shorten the field, or decide the tagline lives somewhere else.
    """
    if len(value) <= limit:
        return value
    detail = (
        f"is {len(value)} characters, longer than the {limit} expected for this field — "
        f"too long for its slot on the page; shorten it at the source"
    )
    acc.warnings.append(f"{where.field} for {where.gtin} {detail}")
    acc.issues.append(
        SourceIssue(
            gtin=where.gtin,
            field=where.field,
            source=where.source,
            issue="value_too_long",
            value=value,
            detail=detail,
        )
    )
    _log.warning("%s for %s %s", where.field, where.gtin, detail)
    return value


def source_label(src: GdsnSource) -> str:
    """Name a field the way the *source system* does, for the report (§SourceIssue).

    Public because ``lib.generator`` needs the same words when it reports a value it filled:
    the operator searches MyGS1 by the attribute, and two spellings of one attribute in one
    report is two things to learn.
    """
    return f"{src.sheet} attr {src.attribute}" if src.attribute else src.sheet


#: Letter patterns near-absent from the *other* language, so their presence in a value declared
#: for one language is a strong hint the source carries the wrong-language text. Deliberately
#: conservative (few, unambiguous patterns) to keep this a low-noise "worth a glance" signal
#: rather than a real language detector: Dutch ``aa``/``uu``/``ij`` essentially never occur in
#: French, and the French cedilla/grave/circumflex essentially never occur in native Dutch.
_DUTCH_HALLMARKS: Final = ("aa", "uu", "ij")
_FRENCH_HALLMARKS: Final = ("ç", "è", "ê", "à", "â", "ù", "û", "œ")


def suspect_language(value: str, lang: str) -> str | None:
    """Return the language ``value`` looks like when it isn't ``lang`` (else ``None``).

    **Public because two consumers must agree.** The report flags these values (§3b) and
    :mod:`lib.generator` now *acts* on them — a slot holding another language's text is treated as
    a language the feed does not carry, so the producer fills it from the one that does. A second
    reading of "is this the wrong language?" would let the report name a value the generator
    quietly left alone, or the reverse.

    That raises the cost of a false positive from a glance to a replaced value, which is why the
    patterns below stay deliberately few. The replacement is not silent: it lands in the report's
    "translated to fill a language gap" section for the operator to paste into MyGS1, beside the
    §3b flag that names the original.
    """
    folded = value.casefold()
    if lang == "fr" and any(h in folded for h in _DUTCH_HALLMARKS):
        return "Dutch"
    if lang == "nl" and any(h in folded for h in _FRENCH_HALLMARKS):
        return "French"
    return None


def _check_field_language(
    values: dict[str, str], field_base: str, source: str, gtin: str
) -> list[SourceIssue]:
    """Flag a localised value carrying hallmarks of a *different* language (heuristic; §4.2).

    A "worth a glance" observation — the value is passed through unchanged, like every other
    source finding. Catches the common case of a translation slot left in the source language
    (e.g. a French ``product_name`` that still reads ``Schoonmaakdoek``). Only ``nl`` and ``fr``
    exist for this client; other languages are skipped.
    """
    issues: list[SourceIssue] = []
    for lang, value in values.items():
        suspect = suspect_language(value, lang)
        if suspect is None:
            continue
        detail = (
            f"the {lang} value {value!r} reads like {suspect} — the source may carry "
            f"{suspect} text in the {lang} slot; check the translation at the source"
        )
        issues.append(
            SourceIssue(
                gtin=gtin,
                field=f"{field_base}.{lang}",
                source=source,
                issue="value_wrong_language",
                value=value,
                detail=detail,
            )
        )
    return issues


def _apply_checks(value: str, src: GdsnSource, field: str, gtin: str, acc: _Accumulator) -> str:
    """Apply the configured source-value expectations, in order."""
    where = _Where(field=field, source=source_label(src), gtin=gtin)
    if src.strip_prefix:
        value = _strip_prefix(value, src.strip_prefix, where, acc)
    if src.max_length:
        # After stripping: the prefix is not part of what renders in the slot.
        value = _check_length(value, src.max_length, where, acc)
    return value


def _localised_picker(
    sheet: GdsnSheet, gtin: str, attribute: str, lang: str
) -> Callable[[str], str | None]:
    """A single-market picker for one language, so :func:`_pick_ranked` sees one arg.

    Binds ``lang`` in a fresh scope per call, which both keeps the returned callable
    single-argument (the shape ``_pick_ranked`` expects) and avoids the classic
    loop-variable capture bug when this is called once per language.
    """
    return lambda market: sheet.pick_localised(gtin, market, attribute, lang)


def _multivalue_picker(
    sheet: GdsnSheet, gtin: str, attribute: str, lang: str
) -> Callable[[str], str | None]:
    """A single-market picker that joins every repeated slot's value with a newline.

    Returns ``None`` when the market carries no slot, so :func:`_pick_ranked` skips it and ranks
    to the next market — matching the single-value picker's contract.
    """

    def pick(market: str) -> str | None:
        values = sheet.pick_localised_all(gtin, market, attribute, lang)
        return "\n".join(values) if values else None

    return pick


def _scalar_picker(sheet: GdsnSheet, gtin: str, src: GdsnSource) -> Callable[[str], str | None]:
    """The single- or every-slot picker for a language-agnostic source, per its ``multivalue``.

    One function because both language-agnostic resolvers — the mapped field and the pass-through
    extra — need the same choice, and letting them each spell it out is how ``multivalue`` came to
    mean one thing on one path and nothing on the others.
    """
    if src.multivalue:
        return _multivalue_scalar_picker(sheet, gtin, src.attribute, src.with_unit)
    return lambda market: sheet.pick_scalar(gtin, market, src.attribute, src.with_unit)


def _multivalue_scalar_picker(
    sheet: GdsnSheet, gtin: str, attribute: str, with_unit: bool
) -> Callable[[str], str | None]:
    """A language-agnostic picker that joins every repeated slot with :data:`SCALAR_SEPARATOR`.

    The scalar counterpart of :func:`_multivalue_picker`. Returns ``None`` when the market carries
    no slot, so :func:`_pick_ranked` skips it and ranks on — matching the single-value contract.
    """

    def pick(market: str) -> str | None:
        values = sheet.pick_scalar_all(gtin, market, attribute, with_unit)
        return SCALAR_SEPARATOR.join(values) if values else None

    return pick


def _pick_ranked(
    pick: Callable[[str], str | None], market_priority: list[str]
) -> tuple[str | None, dict[str, str]]:
    """Walk markets in priority order, collecting each one's non-blank value.

    Returns ``(chosen, per_market)`` where ``chosen`` is the first market's value in
    priority order (``None`` if every market is blank) and ``per_market`` maps each market
    that supplied a value to it — the raw material for both the ranked choice and the
    cross-market inconsistency check, gathered in one pass because you cannot rank without
    seeing every candidate.
    """
    per_market: dict[str, str] = {}
    chosen: str | None = None
    for market in market_priority:
        value = pick(market)
        if value is not None:
            per_market[market] = value
            if chosen is None:
                chosen = value
    return chosen, per_market


def _report_inconsistency(per_market: dict[str, str], where: _Where, acc: _Accumulator) -> None:
    """Report when 2+ markets carry different non-blank values for one field/language (§6).

    The datapool is authoritative, so a disagreement is not the tool's to resolve — it
    picks the ranked winner and reports the conflict for a human to reconcile in MyGS1.

    Case- and whitespace-only differences are *not* reported: ``"voegstrijker"`` vs
    ``"Voegstrijker"`` is not a content disagreement, and the page CSS uppercases the title
    regardless — flagging it buries the substantive conflicts (``"toilettas"`` vs
    ``"Cosmetic Bag"``, ``"5 H87"`` vs ``"1 H87"``) in noise. Accents survive casefolding,
    so a missing diacritic (``"Désherbant"`` vs ``"Desherbant"``) is still a real conflict.
    """
    if len({v.strip().casefold() for v in per_market.values()}) <= 1:
        return
    chosen = next(iter(per_market.values()))
    pairs = "; ".join(f"{market}={value!r}" for market, value in per_market.items())
    detail = (
        f"differs across target markets ({pairs}) — same field, same language, different "
        f"text; the tool used the highest-ranked ({chosen!r}). Decide which market is "
        f"authoritative and align them at the source"
    )
    acc.warnings.append(f"{where.field} for {where.gtin} {detail}")
    acc.issues.append(
        SourceIssue(
            gtin=where.gtin,
            field=where.field,
            source=where.source,
            issue="value_inconsistent_across_markets",
            value=chosen,
            detail=detail,
            market_values=tuple(per_market.items()),
        )
    )
    _log.warning("%s for %s %s", where.field, where.gtin, detail)


def _report_blank(where: _Where, acc: _Accumulator) -> None:
    """Report a published field with no value in any market that describes the product (§6).

    Not raised for a product that simply has no row in a market — that is not a gap
    (:func:`_pick_ranked` only ever sees markets, and a product absent from all of them is
    absent from ``records`` entirely). This fires only when the product exists but the slot
    is empty everywhere, which is a hole on the page and, later, the generator's work list.
    """
    detail = "is empty in every target market that carries this product — fill it at the source"
    acc.warnings.append(f"{where.field} for {where.gtin} {detail}")
    acc.issues.append(
        SourceIssue(
            gtin=where.gtin,
            field=where.field,
            source=where.source,
            issue="value_blank",
            value="",
            detail=detail,
        )
    )
    _log.warning("%s for %s %s", where.field, where.gtin, detail)


def _resolve_field(
    ctx: _BuildContext, field: str, src: GdsnSource, gtin: str, acc: _Accumulator
) -> None:
    """Resolve one mapped field for one GTIN into the accumulator, reporting as it goes."""
    sheet = ctx.workbook.get(src.sheet)
    if sheet is None:
        return
    if src.localised:
        _resolve_localised(ctx, sheet, field, src, gtin, acc)
    elif src.primary_file:
        chosen, _ = _pick_ranked(
            lambda market: sheet.pick_primary_file(gtin, market), ctx.market_priority
        )
        if chosen is not None:
            acc.scalars[field] = chosen
        elif src.report_issues:
            _report_blank(_Where(field, source_label(src), gtin), acc)
    else:
        _resolve_scalar(ctx, sheet, field, src, gtin, acc)


def _resolve_localised(  # noqa: PLR0913 — one collaborator per step; bundling hides them
    ctx: _BuildContext, sheet: GdsnSheet, field: str, src: GdsnSource, gtin: str, acc: _Accumulator
) -> None:
    """Resolve a per-language field across ranked markets, reporting blanks and conflicts."""
    values: dict[str, str] = {}
    for lang in ctx.languages:
        picker = (
            _multivalue_picker(sheet, gtin, src.attribute, lang)
            if src.multivalue
            else _localised_picker(sheet, gtin, src.attribute, lang)
        )
        chosen, per_market = _pick_ranked(picker, ctx.market_priority)
        where = _Where(f"{field}.{lang}", source_label(src), gtin)
        if src.report_issues:
            _report_inconsistency(per_market, where, acc)
        if chosen is not None:
            values[lang] = _apply_checks(chosen, src, f"{field}.{lang}", gtin, acc)
        elif src.report_issues:
            _report_blank(where, acc)
    if values:
        acc.localised[field] = values


def _resolve_scalar(  # noqa: PLR0913 — one collaborator per step; bundling hides them
    ctx: _BuildContext, sheet: GdsnSheet, field: str, src: GdsnSource, gtin: str, acc: _Accumulator
) -> None:
    """Resolve a language-agnostic field from the highest-priority market that carries it."""
    chosen, per_market = _pick_ranked(
        _scalar_picker(sheet, gtin, src),
        ctx.market_priority,
    )
    where = _Where(field, source_label(src), gtin)
    if src.report_issues:
        _report_inconsistency(per_market, where, acc)
    if chosen is not None:
        acc.scalars[field] = _apply_checks(chosen, src, field, gtin, acc)
    elif src.report_issues:
        _report_blank(where, acc)


def _resolve_extras(
    ctx: _BuildContext, gdsn_extras: dict[str, GdsnSource], gtin: str, acc: _Accumulator
) -> None:
    """Resolve every pass-through extra for one GTIN into the accumulator.

    Per-language extras land in :attr:`_Accumulator.extras_localised` and language-agnostic ones
    in :attr:`_Accumulator.extras` — one home per value, so no reader downstream has to decide
    which of two dicts is authoritative for a given name.
    """
    for name, src in gdsn_extras.items():
        if src.localised:
            values = _resolve_localised_extra(ctx, src, gtin)
            if values is not None:
                acc.extras_localised[name] = values
            continue
        value = _resolve_extra(ctx, src, gtin)
        if value is not None:
            acc.extras[name] = value


def _resolve_localised_extra(
    ctx: _BuildContext, src: GdsnSource, gtin: str
) -> dict[str, str] | None:
    """Resolve a per-language pass-through extra into every configured language.

    It used to resolve to the default language alone, which silently discarded every other
    language the feed carried — so a French page rendered the Dutch token, and the quality
    report could not tell a genuine translation gap from one the parser had created.

    Extras stay unreported here: they are not page fields, so a blank or a cross-market
    disagreement in one is not a source-fix finding at parse time.
    """
    sheet = ctx.workbook.get(src.sheet)
    if sheet is None:
        return None
    values: dict[str, str] = {}
    for lang in ctx.languages:
        picker = (
            _multivalue_picker(sheet, gtin, src.attribute, lang)
            if src.multivalue
            else _localised_picker(sheet, gtin, src.attribute, lang)
        )
        chosen, _ = _pick_ranked(picker, ctx.market_priority)
        if chosen is not None:
            values[lang] = chosen
    return values or None


def _resolve_extra(ctx: _BuildContext, src: GdsnSource, gtin: str) -> str | None:
    """Resolve a language-agnostic pass-through extra to a single string.

    Per-language extras go through :func:`_resolve_localised_extra` instead.
    """
    sheet = ctx.workbook.get(src.sheet)
    if sheet is None:
        return None
    if src.primary_file:
        chosen, _ = _pick_ranked(
            lambda market: sheet.pick_primary_file(gtin, market), ctx.market_priority
        )
        return chosen
    chosen, _ = _pick_ranked(_scalar_picker(sheet, gtin, src), ctx.market_priority)
    return chosen
