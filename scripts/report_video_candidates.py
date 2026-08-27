"""Emit the video→GTIN candidate report as a spreadsheet the client can work through.

Usage:
    python -m scripts.report_video_candidates [CLIENT_ID] [--top-n N] [--format csv|xlsx]
                                              [--out PATH] [--products PATH]

``CLIENT_ID`` may be omitted when ``clients.yml`` defines exactly one client.

Read-only. One row per (language, video file) over the **union** of the folders under
``media.video_folders`` and the rows in ``media.video_map_path``, so a file nobody has mapped and
a mapping row whose file never arrived are both visible. Each row carries its state, the GTIN the
mapping holds today, the names of the product that GTIN is, and the top ``--top-n`` fuzzy
candidates — each with the value that scored and the field it came from, because on this feed the
winning field is usually a French one holding English (see :mod:`lib.video_candidates`).

This is a report to hand over, not a gate: it never exits non-zero over an unmapped file. The gate
is ``python -m scripts.build_video_map --check``, which is also what writes ``video_map_issues
.json`` for the data-quality report. Filling a mapping in one row at a time has a screen — the
operator shell's **Video mapping**. This exists for the other job: sending the whole backlog to
the client, who is the one who can settle it.

Emits: output/{client_id}/video-map-candidates.{csv,xlsx}
Exit codes:
    0  report written
    1  the products file or the mapping could not be read
    2  config/usage error (bad client id, no media config, no media.video_map_path)
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lib.config import ClientConfig, get_client
from lib.env import load_env
from lib.errors import ConfigError, VideoMapError
from lib.media_video import VideoMap, list_video_files, load_video_map, summarize_video_map
from lib.records import ProductRecord
from lib.video_candidates import CandidateRow, Cell, build_rows, cells, header

_EXIT_OK = 0
_EXIT_ERROR = 1
_EXIT_USAGE = 2

#: Hints per row. Three is what the drafted mapping comments and the shell's screen both offer;
#: a client working through a backlog usually wants more, hence ``--top-n``.
_DEFAULT_TOP_N = 3

#: Widest a column is auto-sized to. A marketing name can run past a screen, and a sheet whose
#: first column is 300 characters wide is worse to work in than one that truncates on screen.
_MAX_COLUMN_WIDTH = 48


def _default_products_path(client_id: str) -> Path:
    """The parsed-products location written by ``scripts/parse_export.py``."""
    return Path("output") / client_id / "data" / "products.json"


def _default_out_path(client_id: str, suffix: str) -> Path:
    """Beside ``data-quality-report.md``, with the rest of the derived artefacts."""
    return Path("output") / client_id / f"video-map-candidates.{suffix}"


def _load_products(path: Path) -> list[ProductRecord]:
    """Read the parsed-products JSON array into ``ProductRecord``s."""
    data = json.loads(path.read_text(encoding="utf-8"))
    return [ProductRecord.model_validate(item) for item in data]


def _folder_files(cfg: ClientConfig) -> dict[str, list[str]]:
    """Return ``{language: [filename]}`` for each configured video folder."""
    assert cfg.media is not None  # guarded by the caller
    return {
        language: [p.name for p in list_video_files(Path(folder))]
        for language, folder in cfg.media.video_folders.items()
    }


def _write_csv(path: Path, columns: list[str], rows: list[list[Cell]]) -> None:
    """Write the grid as UTF-8 CSV with a BOM.

    The BOM is not decoration: without it Excel on Windows opens a UTF-8 CSV as cp1252, and every
    accented product name in this feed — ``Bâton``, ``Rôtissoire`` — arrives mangled in the file
    that goes to the client.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(columns)
        writer.writerows(rows)


def _write_xlsx(path: Path, columns: list[str], rows: list[list[Cell]]) -> None:
    """Write the grid as a worksheet with a frozen, filterable header row."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "video candidates"
    sheet.append(columns)
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    _format_scores(sheet, columns)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _size_columns(sheet, columns, rows)
    workbook.save(path)
    workbook.close()


def _format_scores(sheet: Any, columns: list[str]) -> None:
    """Show every score to two places.

    The values are written as numbers, not text, so the client can sort and filter a backlog by
    how confident the match is — and a bare float renders ``0.5`` beside ``0.83``, which reads as
    two different kinds of measurement rather than one column.
    """
    for index, name in enumerate(columns, start=1):
        if not name.endswith("_score"):
            continue
        for (cell,) in sheet.iter_rows(min_row=2, min_col=index, max_col=index):
            cell.number_format = "0.00"


def _size_columns(sheet: Any, columns: list[str], rows: list[list[Cell]]) -> None:
    """Widen each column to its widest value, capped — 173 rows is a scrolling job either way."""
    for index, name in enumerate(columns):
        longest = max((len(str(row[index])) for row in rows), default=0)
        width = min(max(len(name), longest) + 2, _MAX_COLUMN_WIDTH)
        sheet.column_dimensions[get_column_letter(index + 1)].width = width


def _report(
    cfg: ClientConfig, vmap: VideoMap, rows: list[CandidateRow], files: dict[str, list[str]]
) -> None:
    """Say what was counted, on stderr, in the same terms ``build_video_map --check`` uses.

    The "no files at all" line comes from :class:`~lib.media_video.VideoMapSummary` rather than
    from a ``len(files) == 0`` here, because that condition already has one owner and has already
    been got wrong once by a surface that counted it for itself.
    """
    states = Counter(row.state for row in rows)
    tally = ", ".join(f"{count} {state}" for state, count in sorted(states.items()))
    print(f"{len(rows)} row(s): {tally}", file=sys.stderr)

    if summarize_video_map(vmap, files, cfg.wordpress.languages).no_files_found:
        print(
            "  no video files found — the folders under media.video_folders are empty or not on "
            "this machine, so every mapped row is reported as NOT ON DISK",
            file=sys.stderr,
        )


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="report_video_candidates",
        description="Emit the video→GTIN candidate report for a client to work through.",
    )
    parser.add_argument(
        "client_id",
        nargs="?",
        help="Key under clients: in clients.yml (optional when only one client is defined)",
    )
    parser.add_argument(
        "--top-n",
        type=int,
        default=_DEFAULT_TOP_N,
        help=f"ranked candidates offered per row (default {_DEFAULT_TOP_N})",
    )
    parser.add_argument(
        "--format",
        choices=("xlsx", "csv"),
        default="xlsx",
        help="output format (default xlsx — the client works in a spreadsheet)",
    )
    parser.add_argument(
        "--out", help="output path (default output/{client_id}/video-map-candidates.{format})"
    )
    parser.add_argument(
        "--products", help="Parsed products JSON (default: output/{id}/data/products.json)"
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """Entry point. Returns the process exit code."""
    args = _parse_args(argv)
    if args.top_n < 1:
        print("error: --top-n must be at least 1", file=sys.stderr)
        return _EXIT_USAGE

    try:
        cfg = get_client(args.client_id)
    except ConfigError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return _EXIT_USAGE

    if cfg.media is None or not cfg.media.video_map_path:
        print(
            f"client {cfg.client_id!r} has no media.video_map_path — there is no video mapping "
            f"to report on",
            file=sys.stderr,
        )
        return _EXIT_USAGE

    products_path = Path(args.products) if args.products else _default_products_path(cfg.client_id)
    try:
        products = _load_products(products_path)
        vmap = load_video_map(Path(cfg.media.video_map_path))
    except (FileNotFoundError, json.JSONDecodeError, VideoMapError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return _EXIT_ERROR

    languages = cfg.wordpress.languages
    files = _folder_files(cfg)
    rows = build_rows(vmap, files, products, languages, top_n=args.top_n)

    columns = header(languages, args.top_n)
    grid = [cells(row, languages, args.top_n) for row in rows]
    out = Path(args.out) if args.out else _default_out_path(cfg.client_id, args.format)
    if args.format == "csv":
        _write_csv(out, columns, grid)
    else:
        _write_xlsx(out, columns, grid)

    print(f"Wrote {out} ({len(rows)} row(s), {len(columns)} column(s))", file=sys.stderr)
    _report(cfg, vmap, rows, files)
    return _EXIT_OK


if __name__ == "__main__":
    load_env()
    raise SystemExit(main())
