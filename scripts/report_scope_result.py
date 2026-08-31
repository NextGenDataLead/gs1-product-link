"""Emit the per-run result sheet: the operator's scope list with what the run did to each row.

Usage:
    python -m scripts.report_scope_result [CLIENT_ID] [--run PATH] [--plan PATH] [--no-plan]
                                          [--list PATH] [--products PATH] [--out PATH]

``CLIENT_ID`` may be omitted when ``clients.yml`` defines exactly one client.

Read-only, and written **after** a run rather than during one. It is a report, not a control file:
nothing here decides what a later run does, which is the whole difference from the design where the
scope list grows a status column and the run reads it back. See ``lib/process_list.py`` for what
that cost the last time it was tried.

Three sheets, one workbook:

* ``scope``   — one row per SKU: the operator's own columns, then ``in_scope``, ``result``, and
                ``status``/``page``/``detail`` per language.
* ``units``   — one row per (GTIN, language) from the run log and the plan's holds, uninterpreted.
                Where "nl published, fr failed" survives.
* ``legend``  — what each value means, so the file can be forwarded without a covering email.

``--run`` defaults to the newest log in ``output/{client_id}/runs`` **by modification time, not by
name**: a same-second second run is written as ``{ts}-1.jsonl``, which sorts *before* ``{ts}.jsonl``
because ``-`` precedes ``.``.

Rows come from the **uploaded** list (``process-list.source.xlsx``) when it is there, so a row the
operator deselected is reported as deselected rather than being missing from their own report. With
no archive it falls back to the control file and says so — the report is then about the rows that
ran, and the deselected ones cannot be named because nothing recorded them.

Emits: output/{client_id}/runs/{run stem}-scope.xlsx
Exit codes:
    0  report written
    1  the run log, the scope list or the products file could not be read
    2  config/usage error (bad client id, no process_list block, no run log to report on)
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from lib.config import ClientConfig, ProcessListConfig, get_client
from lib.env import load_env
from lib.errors import ConfigError, ProcessListError
from lib.process_list import ProcessListSheet, read_process_list
from lib.records import Plan, ProductRecord, RunOutcome
from lib.scope_report import build_rows, legend_grid, scope_grid, units_grid

_EXIT_OK = 0
_EXIT_ERROR = 1
_EXIT_USAGE = 2

#: Widest a column is auto-sized to — an error string runs to a paragraph, and a sheet whose first
#: column is 300 characters wide is worse to work in than one that truncates on screen.
_MAX_COLUMN_WIDTH = 60


def _newest_run(client_id: str) -> Path | None:
    """The most recent run log, by mtime. See the module docstring for why not by name."""
    try:
        paths = sorted(
            (Path("output") / client_id / "runs").glob("*.jsonl"), key=lambda p: p.stat().st_mtime
        )
    except OSError:
        return None
    return paths[-1] if paths else None


def _load_outcomes(path: Path) -> tuple[list[RunOutcome], int]:
    """Read a run log, keeping the rows that parse and counting the ones that do not.

    A truncated final line is normal for a run killed mid-write, and discarding the whole file over
    it would throw away the record exactly when it matters most — which is also when this report
    is most likely to be asked for.
    """
    outcomes: list[RunOutcome] = []
    unreadable = 0
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            outcomes.append(RunOutcome.model_validate_json(line))
        except ValueError:
            unreadable += 1
    return outcomes, unreadable


def _load_plan(path: Path) -> Plan | None:
    """The plan, for its holds. ``None`` when it is absent or will not validate."""
    try:
        return Plan.model_validate_json(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None


def _plan_is_for(plan: Plan, outcomes: list[RunOutcome]) -> bool:
    """Whether this plan plausibly belongs to this run.

    ``plan.json`` is overwritten by every ``run_plan``, so for anything but the newest run it is
    somebody else's document. A plan generated *after* the run it is being read beside describes
    work that run never saw, and its holds would be reported as that run's — so it is refused and
    said out loud rather than quietly used.
    """
    return not outcomes or plan.generated_at <= max(outcome.ts for outcome in outcomes)


def _load_products(path: Path) -> list[ProductRecord]:
    data = json.loads(path.read_text(encoding="utf-8"))
    return [ProductRecord.model_validate(item) for item in data]


def _scope_sheets(cfg: ClientConfig) -> tuple[ProcessListSheet, ProcessListSheet, bool]:
    """The uploaded list and the control list. The flag says whether the archive was really there.

    Falls back to the control file for both, so a client who has never used the shell's upload
    still gets a report — one that cannot name the deselected rows, because nothing recorded them.
    """
    assert cfg.process_list is not None  # guarded by the caller
    control = read_process_list(cfg.process_list)
    archive = control.path.with_suffix(f".source{control.path.suffix}")
    if not archive.exists():
        return control, control, False
    uploaded = read_process_list(
        ProcessListConfig(path=str(archive), gtin_column=cfg.process_list.gtin_column)
    )
    return uploaded, control, True


def _write_xlsx(path: Path, sheets: list[tuple[str, list[str], list[list[str]]]]) -> None:
    """Write the workbook, each sheet with a frozen, filterable header row."""
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for title, columns, rows in sheets:
        sheet = workbook.create_sheet(title)
        sheet.append(columns)
        for row in rows:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        _size_columns(sheet, columns, rows)
    workbook.save(path)
    workbook.close()


def _size_columns(sheet: Any, columns: list[str], rows: list[list[str]]) -> None:
    """Widen each column to its widest value, capped."""
    for index, name in enumerate(columns):
        longest = max((len(str(row[index])) for row in rows if index < len(row)), default=0)
        width = min(max(len(name), longest) + 2, _MAX_COLUMN_WIDTH)
        sheet.column_dimensions[get_column_letter(index + 1)].width = width


def _report(rows: list[Any], unreadable: int, archived: bool, plan_used: bool) -> None:
    """Say what was counted, on stderr, naming every reason a number is lower than it looks."""
    results = Counter(row.result for row in rows)
    tally = ", ".join(f"{count} {name}" for name, count in sorted(results.items()))
    print(f"{len(rows)} SKU(s): {tally}", file=sys.stderr)
    if unreadable:
        print(
            f"  {unreadable} line(s) of the run log did not parse and are not in this report — "
            "usually a run killed mid-write",
            file=sys.stderr,
        )
    if not archived:
        print(
            "  no uploaded list archived beside the control file, so the rows reported are the "
            "ones that ran; any row deselected before the run cannot be named",
            file=sys.stderr,
        )
    if not plan_used:
        print(
            "  no plan read, so a unit the plan held reads `not run` rather than `held` — pass "
            "--plan to point at the plan this run was executed from",
            file=sys.stderr,
        )


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="report_scope_result",
        description="Emit the per-run result sheet for a client's scope list.",
    )
    parser.add_argument(
        "client_id",
        nargs="?",
        help="Key under clients: in clients.yml (optional when only one client is defined)",
    )
    parser.add_argument("--run", help="Run log (default: the newest output/{id}/runs/*.jsonl)")
    parser.add_argument("--plan", help="Plan to read holds from (default: output/{id}/plan.json)")
    parser.add_argument(
        "--no-plan",
        action="store_true",
        help="Do not read a plan; held units then read `not run`",
    )
    parser.add_argument("--list", help="Scope list to report on (default: the client's, uploaded)")
    parser.add_argument(
        "--products", help="Parsed products JSON (default: output/{id}/data/products.json)"
    )
    parser.add_argument(
        "--out", help="Output path (default: beside the run log, {stem}-scope.xlsx)"
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:  # noqa: PLR0911, PLR0912 — one exit per failure
    """Entry point. Returns the process exit code."""
    args = _parse_args(argv)
    try:
        cfg = get_client(args.client_id)
    except ConfigError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return _EXIT_USAGE

    if cfg.process_list is None:
        print(
            f"client {cfg.client_id!r} has no process_list block — there is no scope list to "
            f"report on",
            file=sys.stderr,
        )
        return _EXIT_USAGE

    run_path = Path(args.run) if args.run else _newest_run(cfg.client_id)
    if run_path is None:
        print(
            f"no run log under output/{cfg.client_id}/runs — there is nothing to report on yet",
            file=sys.stderr,
        )
        return _EXIT_USAGE

    products_path = (
        Path(args.products)
        if args.products
        else Path("output") / cfg.client_id / "data" / "products.json"
    )
    try:
        outcomes, unreadable = _load_outcomes(run_path)
        products = _load_products(products_path)
        if args.list:
            listed = ProcessListConfig(path=args.list, gtin_column=cfg.process_list.gtin_column)
            uploaded = control = read_process_list(listed)
            archived = True
        else:
            uploaded, control, archived = _scope_sheets(cfg)
    except (OSError, json.JSONDecodeError, ValueError, ProcessListError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return _EXIT_ERROR

    plan = _read_plan(args, cfg.client_id, outcomes)
    rows = build_rows(
        uploaded,
        selected=control.listed_gtins(),
        exported={product.gtin14 for product in products},
        outcomes=outcomes,
        skipped=plan.skipped if plan else [],
        languages=cfg.wordpress.languages,
    )

    columns, grid = scope_grid(uploaded, rows, cfg.wordpress.languages)
    out = Path(args.out) if args.out else run_path.with_name(f"{run_path.stem}-scope.xlsx")
    _write_xlsx(
        out,
        [
            ("scope", columns, grid),
            ("units", *units_grid(outcomes, plan.skipped if plan else [])),
            ("legend", *legend_grid()),
        ],
    )
    print(f"Wrote {out} for {run_path.name}", file=sys.stderr)
    _report(rows, unreadable, archived, plan is not None)
    return _EXIT_OK


def _read_plan(args: argparse.Namespace, client_id: str, outcomes: list[RunOutcome]) -> Plan | None:
    """The plan whose holds this report names, or ``None`` with a reason on stderr."""
    if args.no_plan:
        return None
    path = Path(args.plan) if args.plan else Path("output") / client_id / "plan.json"
    plan = _load_plan(path)
    if plan is None:
        if args.plan:
            print(f"warning: {path} could not be read as a plan", file=sys.stderr)
        return None
    if not _plan_is_for(plan, outcomes):
        print(
            f"warning: {path} was generated after this run, so it is a later run's plan — its "
            f"holds are not reported. Pass --plan to point at the right one, or --no-plan.",
            file=sys.stderr,
        )
        return None
    return plan


if __name__ == "__main__":
    load_env()
    raise SystemExit(main())
